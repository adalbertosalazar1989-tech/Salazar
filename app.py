from __future__ import annotations

import html
import io
import json
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import Element, SubElement, tostring

from cryptography.hazmat.primitives.serialization import pkcs12
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, PlainTextResponse, Response
from openpyxl import load_workbook

APP_NAME = "SAENG Software SST V2"
APP_VERSION = "2.0.0-local"
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("SAENG_DB_PATH", str(BASE_DIR / "data" / "saeng_v2.db")))
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

app = FastAPI(title=APP_NAME, version=APP_VERSION)

CSS = """
:root{--navy:#061b33;--navy2:#0b2949;--gold:#c9a24a;--bg:#f4f7fb;--text:#142033;--muted:#66758a;--danger:#9f2d2d;--ok:#13734b}
*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--text)}
a{color:inherit;text-decoration:none}.shell{min-height:100vh;display:grid;grid-template-columns:260px 1fr}.side{background:linear-gradient(180deg,var(--navy),var(--navy2));color:white;padding:24px 18px;position:sticky;top:0;height:100vh}.brand{display:flex;gap:12px;align-items:center;margin-bottom:28px}.mark{width:48px;height:48px;border:2px solid var(--gold);border-radius:12px;display:grid;place-items:center;color:var(--gold);font-weight:800}.brand strong{display:block}.brand small{color:#b9c6d6}.nav a{display:block;padding:11px 12px;border-radius:9px;margin:4px 0;color:#d7e0ea}.nav a:hover{background:#ffffff14;color:white}.main{padding:28px}.top{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;margin-bottom:20px}.top h1{margin:0;font-size:28px}.top p{margin:5px 0 0;color:var(--muted)}.badge{background:#fff5d8;color:#765a14;border:1px solid #e6c86f;padding:8px 12px;border-radius:999px;font-size:13px}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:16px}.card{background:white;border:1px solid #dfe6ef;border-radius:14px;padding:18px;box-shadow:0 5px 18px #10243a0d}.metric b{font-size:30px;display:block;color:var(--navy)}.metric span{color:var(--muted)}table{width:100%;border-collapse:collapse;background:white;border-radius:12px;overflow:hidden}th,td{padding:11px 10px;border-bottom:1px solid #e7edf3;text-align:left;vertical-align:top;font-size:14px}th{background:#eef3f8;color:#2a3d52}form{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.full{grid-column:1/-1}label{font-size:13px;font-weight:700;color:#40536a}input,select,textarea{width:100%;padding:10px 11px;border:1px solid #cfd9e5;border-radius:8px;background:white;font:inherit;margin-top:5px}textarea{min-height:92px}.btn{display:inline-flex;justify-content:center;align-items:center;border:0;border-radius:9px;padding:11px 15px;background:var(--navy);color:white;font-weight:700;cursor:pointer}.btn.gold{background:var(--gold);color:#152036}.btn.light{background:#e8eef5;color:#17304b}.alert{padding:12px 14px;border-radius:10px;margin:12px 0;background:#eef5ff;border:1px solid #b9d1ee}.alert.error{background:#fff0f0;border-color:#efb5b5;color:var(--danger)}.alert.ok{background:#effaf4;border-color:#a8d9bd;color:var(--ok)}.pill{display:inline-block;padding:4px 8px;border-radius:999px;background:#edf2f7;font-size:12px}.section{margin-top:20px}.muted{color:var(--muted)}.hero{background:linear-gradient(135deg,var(--navy),#0c3962);color:white;border-radius:18px;padding:28px;display:grid;grid-template-columns:1.25fr .75fr;gap:22px}.hero h2{font-size:34px;margin:0 0 10px}.hero p{color:#d3deeb}.hero .box{background:white;color:var(--text);border-radius:14px;padding:18px}.hero .box h3{margin-top:0}.footer{color:var(--muted);font-size:12px;margin-top:30px}@media(max-width:950px){.shell{grid-template-columns:1fr}.side{position:relative;height:auto}.grid{grid-template-columns:repeat(2,1fr)}.hero{grid-template-columns:1fr}}@media(max-width:620px){.main{padding:16px}.grid,form{grid-template-columns:1fr}.full{grid-column:auto}.top{display:block}}
"""

NAV = [
    ("/", "Painel"), ("/companies", "Empresas"), ("/workers", "Trabalhadores"),
    ("/risks", "Riscos e agentes"), ("/exams", "Exames"), ("/events", "Eventos SST"),
    ("/imports", "Importacao XLSX"), ("/certificate", "Certificado A1"),
    ("/readiness", "Prontidao"), ("/audit", "Auditoria"),
]


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def utcnow() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS companies(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                legal_name TEXT NOT NULL,
                cnpj TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'ativa',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS workers(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                name TEXT NOT NULL,
                cpf TEXT NOT NULL,
                registration TEXT NOT NULL,
                role_name TEXT,
                cbo TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, registration)
            );
            CREATE TABLE IF NOT EXISTS risks(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                worker_id INTEGER REFERENCES workers(id),
                category TEXT NOT NULL,
                product_name TEXT,
                substance TEXT,
                cas_number TEXT,
                table24_code TEXT,
                exposure_form TEXT,
                exposure_route TEXT,
                frequency TEXT,
                evaluation_type TEXT NOT NULL,
                intensity TEXT,
                unit TEXT,
                technique TEXT,
                epi TEXT,
                epc TEXT,
                table27_code TEXT,
                exam_name TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS exams(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                worker_id INTEGER NOT NULL REFERENCES workers(id),
                aso_type TEXT NOT NULL,
                aso_date TEXT NOT NULL,
                result TEXT NOT NULL,
                table27_code TEXT,
                procedure_name TEXT,
                physician TEXT,
                crm TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS events(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                worker_id INTEGER REFERENCES workers(id),
                event_type TEXT NOT NULL,
                status TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                xml_text TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS audit(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                entity TEXT NOT NULL,
                entity_id TEXT,
                detail TEXT,
                created_at TEXT NOT NULL
            );
            """
        )


def audit(action: str, entity: str, entity_id: Any = None, detail: str = "") -> None:
    with db() as conn:
        conn.execute(
            "INSERT INTO audit(action,entity,entity_id,detail,created_at) VALUES(?,?,?,?,?)",
            (action, entity, str(entity_id or ""), detail, utcnow()),
        )


def rows(sql: str, params: tuple[Any, ...] = ()) -> list[sqlite3.Row]:
    with db() as conn:
        return list(conn.execute(sql, params).fetchall())


def one(sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute(sql, params).fetchone()


def page(title: str, body: str, subtitle: str = "Gestao local de eSocial SST") -> HTMLResponse:
    nav = "".join(f'<a href="{url}">{esc(label)}</a>' for url, label in NAV)
    content = f"""<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{esc(title)} - {APP_NAME}</title><style>{CSS}</style></head><body><div class='shell'><aside class='side'><div class='brand'><div class='mark'>SS</div><div><strong>SAENG</strong><small>Software SST V2</small></div></div><nav class='nav'>{nav}</nav></aside><main class='main'><header class='top'><div><h1>{esc(title)}</h1><p>{esc(subtitle)}</p></div><span class='badge'>Ambiente local · sem efeito juridico</span></header>{body}<div class='footer'>SAENG Software SST V2 {APP_VERSION} · Execucao restrita a 127.0.0.1</div></main></div></body></html>"""
    return HTMLResponse(content)


def options(items: list[sqlite3.Row], label_fields: tuple[str, ...], selected: Any = None) -> str:
    out = ["<option value=''>Selecione</option>"]
    for item in items:
        label = " - ".join(str(item[field]) for field in label_fields if item[field] is not None)
        sel = " selected" if str(item["id"]) == str(selected) else ""
        out.append(f"<option value='{item['id']}'{sel}>{esc(label)}</option>")
    return "".join(out)


def build_xml(event_type: str, company: sqlite3.Row, worker: sqlite3.Row | None, payload: dict[str, Any]) -> str:
    root = Element("eSocial")
    event = SubElement(root, event_type.replace("-", ""))
    ide = SubElement(event, "ideEvento")
    SubElement(ide, "tpAmb").text = "2"
    SubElement(ide, "procEmi").text = "1"
    SubElement(ide, "verProc").text = APP_VERSION
    employer = SubElement(event, "ideEmpregador")
    SubElement(employer, "tpInsc").text = "1"
    SubElement(employer, "nrInsc").text = "".join(ch for ch in company["cnpj"] if ch.isdigit())[:8]
    if worker:
        employee = SubElement(event, "ideVinculo")
        SubElement(employee, "cpfTrab").text = "".join(ch for ch in worker["cpf"] if ch.isdigit())
        SubElement(employee, "matricula").text = worker["registration"]
    details = SubElement(event, "dadosLocais")
    for key, value in payload.items():
        child = SubElement(details, str(key))
        child.text = str(value)
    return tostring(root, encoding="unicode")


init_db()


@app.get("/health")
def health() -> dict[str, Any]:
    return {"status": "ok", "app": APP_NAME, "version": APP_VERSION, "database": str(DB_PATH)}


@app.get("/", response_class=HTMLResponse)
def dashboard() -> HTMLResponse:
    counts = {
        "Empresas": one("SELECT COUNT(*) n FROM companies")["n"],
        "Trabalhadores": one("SELECT COUNT(*) n FROM workers")["n"],
        "Riscos": one("SELECT COUNT(*) n FROM risks")["n"],
        "Eventos": one("SELECT COUNT(*) n FROM events")["n"],
    }
    metrics = "".join(f"<div class='card metric'><b>{value}</b><span>{esc(key)}</span></div>" for key, value in counts.items())
    body = f"""<section class='hero'><div><h2>Gestao inteligente de eSocial SST.</h2><p>Empresas, trabalhadores, riscos, agentes quimicos, exames, eventos e auditoria em uma unica base local.</p><div class='alert'>A V2 gera e valida dados localmente. A transmissao oficial permanece bloqueada ate configuracao dos servicos oficiais, XSDs e assinatura.</div></div><div class='box'><h3>Inicio rapido</h3><ol><li>Cadastre a empresa.</li><li>Cadastre trabalhadores.</li><li>Importe ou registre riscos.</li><li>Cadastre exames.</li><li>Gere o evento local.</li></ol><a class='btn gold' href='/companies'>Comecar</a></div></section><section class='section grid'>{metrics}</section>"""
    return page("Painel executivo", body)


@app.get("/companies", response_class=HTMLResponse)
def companies(message: str = "") -> HTMLResponse:
    data = rows("SELECT * FROM companies ORDER BY id DESC")
    trs = "".join(f"<tr><td>{r['id']}</td><td>{esc(r['legal_name'])}</td><td>{esc(r['cnpj'])}</td><td><span class='pill'>{esc(r['status'])}</span></td></tr>" for r in data) or "<tr><td colspan='4'>Nenhuma empresa cadastrada.</td></tr>"
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else ""
    body = f"""{notice}<div class='card'><h3>Nova empresa</h3><form method='post'><label>Razao social<input name='legal_name' required></label><label>CNPJ<input name='cnpj' required></label><label>Status<select name='status'><option>ativa</option><option>suspensa</option><option>arquivada</option></select></label><div><button class='btn' type='submit'>Salvar empresa</button></div></form></div><div class='section'><table><thead><tr><th>ID</th><th>Razao social</th><th>CNPJ</th><th>Status</th></tr></thead><tbody>{trs}</tbody></table></div>"""
    return page("Empresas", body)


@app.post("/companies", response_class=HTMLResponse)
def companies_create(legal_name: str = Form(...), cnpj: str = Form(...), status: str = Form("ativa")) -> HTMLResponse:
    digits = "".join(ch for ch in cnpj if ch.isdigit())
    if len(digits) != 14:
        return page("Empresas", "<div class='alert error'>CNPJ deve possuir 14 digitos.</div>")
    try:
        with db() as conn:
            cur = conn.execute("INSERT INTO companies(legal_name,cnpj,status,created_at) VALUES(?,?,?,?)", (legal_name.strip(), digits, status, utcnow()))
            company_id = cur.lastrowid
        audit("CREATE", "company", company_id, legal_name)
    except sqlite3.IntegrityError:
        return page("Empresas", "<div class='alert error'>CNPJ ja cadastrado.</div>")
    return companies("Empresa cadastrada com sucesso.")


@app.get("/workers", response_class=HTMLResponse)
def workers(message: str = "") -> HTMLResponse:
    companies_data = rows("SELECT id,legal_name,cnpj FROM companies WHERE status='ativa' ORDER BY legal_name")
    data = rows("SELECT w.*,c.legal_name FROM workers w JOIN companies c ON c.id=w.company_id ORDER BY w.id DESC")
    trs = "".join(f"<tr><td>{esc(r['name'])}</td><td>{esc(r['cpf'])}</td><td>{esc(r['registration'])}</td><td>{esc(r['role_name'])}</td><td>{esc(r['legal_name'])}</td></tr>" for r in data) or "<tr><td colspan='5'>Nenhum trabalhador cadastrado.</td></tr>"
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else ""
    body = f"""{notice}<div class='card'><h3>Novo trabalhador</h3><form method='post'><label>Empresa<select name='company_id' required>{options(companies_data,('legal_name','cnpj'))}</select></label><label>Nome<input name='name' required></label><label>CPF<input name='cpf' required></label><label>Matricula<input name='registration' required></label><label>Funcao<input name='role_name'></label><label>CBO<input name='cbo'></label><div class='full'><button class='btn' type='submit'>Salvar trabalhador</button></div></form></div><div class='section'><table><thead><tr><th>Nome</th><th>CPF</th><th>Matricula</th><th>Funcao</th><th>Empresa</th></tr></thead><tbody>{trs}</tbody></table></div>"""
    return page("Trabalhadores", body)


@app.post("/workers", response_class=HTMLResponse)
def workers_create(company_id: int = Form(...), name: str = Form(...), cpf: str = Form(...), registration: str = Form(...), role_name: str = Form(""), cbo: str = Form("")) -> HTMLResponse:
    digits = "".join(ch for ch in cpf if ch.isdigit())
    if len(digits) != 11:
        return page("Trabalhadores", "<div class='alert error'>CPF deve possuir 11 digitos.</div>")
    try:
        with db() as conn:
            cur = conn.execute("INSERT INTO workers(company_id,name,cpf,registration,role_name,cbo,created_at) VALUES(?,?,?,?,?,?,?)", (company_id, name.strip(), digits, registration.strip(), role_name.strip(), cbo.strip(), utcnow()))
            worker_id = cur.lastrowid
        audit("CREATE", "worker", worker_id, name)
    except sqlite3.IntegrityError:
        return page("Trabalhadores", "<div class='alert error'>Matricula ja cadastrada para a empresa.</div>")
    return workers("Trabalhador cadastrado com sucesso.")


@app.get("/risks", response_class=HTMLResponse)
def risks(message: str = "", error: str = "") -> HTMLResponse:
    companies_data = rows("SELECT id,legal_name,cnpj FROM companies WHERE status='ativa' ORDER BY legal_name")
    workers_data = rows("SELECT id,name,registration FROM workers ORDER BY name")
    data = rows("SELECT r.*,c.legal_name,w.name worker_name FROM risks r JOIN companies c ON c.id=r.company_id LEFT JOIN workers w ON w.id=r.worker_id ORDER BY r.id DESC LIMIT 200")
    trs = "".join(f"<tr><td>{esc(r['category'])}</td><td>{esc(r['product_name'])}</td><td>{esc(r['substance'])}</td><td>{esc(r['table24_code'])}</td><td>{esc(r['evaluation_type'])}</td><td>{esc(r['worker_name'] or 'GHE/geral')}</td></tr>" for r in data) or "<tr><td colspan='6'>Nenhum risco cadastrado.</td></tr>"
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else (f"<div class='alert error'>{esc(error)}</div>" if error else "")
    body = f"""{notice}<div class='card'><h3>Novo risco ou agente</h3><div class='alert'>Para risco quimico, informe obrigatoriamente a substancia e o codigo da Tabela 24. Nome comercial isolado nao e aceito.</div><form method='post'><label>Empresa<select name='company_id' required>{options(companies_data,('legal_name','cnpj'))}</select></label><label>Trabalhador opcional<select name='worker_id'>{options(workers_data,('name','registration'))}</select></label><label>Categoria<select name='category'><option value='quimico'>Quimico</option><option value='fisico'>Fisico</option><option value='biologico'>Biologico</option><option value='ergonomico'>Ergonomico</option><option value='acidente'>Acidente</option></select></label><label>Produto comercial<input name='product_name'></label><label>Substancia<input name='substance'></label><label>CAS<input name='cas_number'></label><label>Codigo Tabela 24<input name='table24_code' placeholder='Ex.: 01.18.001'></label><label>Forma de exposicao<input name='exposure_form' placeholder='poeira, vapor, fumo...'></label><label>Via de exposicao<input name='exposure_route' placeholder='inalatoria, dermica...'></label><label>Frequencia<input name='frequency'></label><label>Avaliacao<select name='evaluation_type'><option value='qualitativa'>Qualitativa</option><option value='quantitativa'>Quantitativa</option></select></label><label>Intensidade/concentracao<input name='intensity'></label><label>Unidade<input name='unit'></label><label>Tecnica de medicao<input name='technique'></label><label>EPI<input name='epi'></label><label>EPC<input name='epc'></label><label>Codigo Tabela 27<input name='table27_code'></label><label>Exame relacionado<input name='exam_name'></label><div class='full'><button class='btn' type='submit'>Salvar risco</button></div></form></div><div class='section'><table><thead><tr><th>Categoria</th><th>Produto</th><th>Substancia</th><th>Tabela 24</th><th>Avaliacao</th><th>Vinculo</th></tr></thead><tbody>{trs}</tbody></table></div>"""
    return page("Riscos ocupacionais e agentes", body)


@app.post("/risks", response_class=HTMLResponse)
def risks_create(company_id: int = Form(...), worker_id: str = Form(""), category: str = Form(...), product_name: str = Form(""), substance: str = Form(""), cas_number: str = Form(""), table24_code: str = Form(""), exposure_form: str = Form(""), exposure_route: str = Form(""), frequency: str = Form(""), evaluation_type: str = Form(...), intensity: str = Form(""), unit: str = Form(""), technique: str = Form(""), epi: str = Form(""), epc: str = Form(""), table27_code: str = Form(""), exam_name: str = Form("")) -> HTMLResponse:
    if category == "quimico" and (not substance.strip() or not table24_code.strip()):
        return risks(error="Risco quimico exige substancia e codigo da Tabela 24.")
    if evaluation_type == "quantitativa" and (not intensity.strip() or not unit.strip() or not technique.strip()):
        return risks(error="Avaliacao quantitativa exige intensidade, unidade e tecnica de medicao.")
    wid = int(worker_id) if worker_id.strip() else None
    with db() as conn:
        cur = conn.execute(
            """INSERT INTO risks(company_id,worker_id,category,product_name,substance,cas_number,table24_code,exposure_form,exposure_route,frequency,evaluation_type,intensity,unit,technique,epi,epc,table27_code,exam_name,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id,wid,category,product_name.strip(),substance.strip(),cas_number.strip(),table24_code.strip(),exposure_form.strip(),exposure_route.strip(),frequency.strip(),evaluation_type,intensity.strip(),unit.strip(),technique.strip(),epi.strip(),epc.strip(),table27_code.strip(),exam_name.strip(),utcnow()),
        )
        risk_id = cur.lastrowid
    audit("CREATE", "risk", risk_id, f"{category}:{substance or product_name}")
    return risks(message="Risco cadastrado e validado localmente.")


@app.get("/exams", response_class=HTMLResponse)
def exams(message: str = "") -> HTMLResponse:
    companies_data = rows("SELECT id,legal_name,cnpj FROM companies WHERE status='ativa' ORDER BY legal_name")
    workers_data = rows("SELECT id,name,registration FROM workers ORDER BY name")
    data = rows("SELECT e.*,w.name worker_name,c.legal_name FROM exams e JOIN workers w ON w.id=e.worker_id JOIN companies c ON c.id=e.company_id ORDER BY e.id DESC")
    trs = "".join(f"<tr><td>{esc(r['worker_name'])}</td><td>{esc(r['aso_type'])}</td><td>{esc(r['aso_date'])}</td><td>{esc(r['result'])}</td><td>{esc(r['table27_code'])}</td><td>{esc(r['procedure_name'])}</td></tr>" for r in data) or "<tr><td colspan='6'>Nenhum exame cadastrado.</td></tr>"
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else ""
    body = f"""{notice}<div class='card'><h3>Novo exame ocupacional</h3><form method='post'><label>Empresa<select name='company_id' required>{options(companies_data,('legal_name','cnpj'))}</select></label><label>Trabalhador<select name='worker_id' required>{options(workers_data,('name','registration'))}</select></label><label>Tipo ASO<select name='aso_type'><option>admissional</option><option>periodico</option><option>retorno</option><option>mudanca de risco</option><option>demissional</option></select></label><label>Data<input type='date' name='aso_date' required></label><label>Resultado<select name='result'><option>apto</option><option>inapto</option></select></label><label>Codigo Tabela 27<input name='table27_code'></label><label>Procedimento<input name='procedure_name'></label><label>Medico examinador<input name='physician'></label><label>CRM/UF<input name='crm'></label><div class='full'><button class='btn' type='submit'>Salvar exame</button></div></form></div><div class='section'><table><thead><tr><th>Trabalhador</th><th>Tipo</th><th>Data</th><th>Resultado</th><th>Tabela 27</th><th>Procedimento</th></tr></thead><tbody>{trs}</tbody></table></div>"""
    return page("Exames ocupacionais", body)


@app.post("/exams", response_class=HTMLResponse)
def exams_create(company_id: int = Form(...), worker_id: int = Form(...), aso_type: str = Form(...), aso_date: str = Form(...), result: str = Form(...), table27_code: str = Form(""), procedure_name: str = Form(""), physician: str = Form(""), crm: str = Form("")) -> HTMLResponse:
    with db() as conn:
        cur = conn.execute("INSERT INTO exams(company_id,worker_id,aso_type,aso_date,result,table27_code,procedure_name,physician,crm,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)", (company_id,worker_id,aso_type,aso_date,result,table27_code.strip(),procedure_name.strip(),physician.strip(),crm.strip(),utcnow()))
        exam_id = cur.lastrowid
    audit("CREATE", "exam", exam_id, f"{aso_type}:{aso_date}")
    return exams("Exame cadastrado com sucesso.")


@app.get("/events", response_class=HTMLResponse)
def events(message: str = "", error: str = "") -> HTMLResponse:
    companies_data = rows("SELECT id,legal_name,cnpj FROM companies WHERE status='ativa' ORDER BY legal_name")
    workers_data = rows("SELECT id,name,registration FROM workers ORDER BY name")
    data = rows("SELECT e.*,c.legal_name,w.name worker_name FROM events e JOIN companies c ON c.id=e.company_id LEFT JOIN workers w ON w.id=e.worker_id ORDER BY e.id DESC")
    trs = "".join(f"<tr><td>{r['id']}</td><td>{esc(r['event_type'])}</td><td>{esc(r['legal_name'])}</td><td>{esc(r['worker_name'])}</td><td><span class='pill'>{esc(r['status'])}</span></td><td><a href='/events/{r['id']}/xml'>XML</a></td></tr>" for r in data) or "<tr><td colspan='6'>Nenhum evento gerado.</td></tr>"
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else (f"<div class='alert error'>{esc(error)}</div>" if error else "")
    body = f"""{notice}<div class='card'><h3>Gerar evento local</h3><form method='post'><label>Empresa<select name='company_id' required>{options(companies_data,('legal_name','cnpj'))}</select></label><label>Trabalhador<select name='worker_id'>{options(workers_data,('name','registration'))}</select></label><label>Evento<select name='event_type'><option>S-2210</option><option>S-2220</option><option>S-2240</option><option>S-3000</option></select></label><label>Data da ocorrencia<input type='date' name='event_date' required></label><label class='full'>Dados complementares<textarea name='notes' placeholder='Resumo tecnico do evento'></textarea></label><div class='full'><button class='btn' type='submit'>Validar e gerar XML local</button></div></form></div><div class='section'><table><thead><tr><th>ID</th><th>Evento</th><th>Empresa</th><th>Trabalhador</th><th>Status</th><th>Arquivo</th></tr></thead><tbody>{trs}</tbody></table></div>"""
    return page("Eventos SST", body)


@app.post("/events", response_class=HTMLResponse)
def events_create(company_id: int = Form(...), worker_id: str = Form(""), event_type: str = Form(...), event_date: str = Form(...), notes: str = Form("")) -> HTMLResponse:
    if event_type not in {"S-2210", "S-2220", "S-2240", "S-3000"}:
        return events(error="Tipo de evento invalido.")
    company = one("SELECT * FROM companies WHERE id=?", (company_id,))
    if not company:
        return events(error="Empresa inexistente.")
    worker = one("SELECT * FROM workers WHERE id=?", (int(worker_id),)) if worker_id.strip() else None
    if event_type != "S-3000" and not worker:
        return events(error="Selecione um trabalhador para este evento.")
    payload = {"data": event_date, "observacao": notes.strip(), "origem": "SAENG-V2-local"}
    xml_text = build_xml(event_type, company, worker, payload)
    with db() as conn:
        cur = conn.execute("INSERT INTO events(company_id,worker_id,event_type,status,payload_json,xml_text,created_at) VALUES(?,?,?,?,?,?,?)", (company_id, worker["id"] if worker else None, event_type, "validado_local", json.dumps(payload, ensure_ascii=False), xml_text, utcnow()))
        event_id = cur.lastrowid
    audit("CREATE", "event", event_id, event_type)
    return events(message="Evento validado e XML local gerado. Nao transmitido ao eSocial.")


@app.get("/events/{event_id}/xml")
def event_xml(event_id: int) -> Response:
    event = one("SELECT * FROM events WHERE id=?", (event_id,))
    if not event:
        raise HTTPException(status_code=404, detail="Evento nao encontrado")
    headers = {"Content-Disposition": f"attachment; filename={event['event_type']}_{event_id}.xml"}
    return Response(event["xml_text"], media_type="application/xml", headers=headers)


@app.get("/imports", response_class=HTMLResponse)
def imports(message: str = "", error: str = "") -> HTMLResponse:
    notice = f"<div class='alert ok'>{esc(message)}</div>" if message else (f"<div class='alert error'>{esc(error)}</div>" if error else "")
    body = f"""{notice}<div class='card'><h3>Importar planilha mestra</h3><p class='muted'>A primeira linha deve conter cabecalhos. O sistema reconhece variacoes de Empresa, CNPJ, Categoria, Produto, Substancia, CAS, Tabela 24, Avaliacao, Intensidade, Unidade, Tecnica, EPI, EPC, Tabela 27 e Exame.</p><form method='post' enctype='multipart/form-data'><label class='full'>Arquivo XLSX<input type='file' name='file' accept='.xlsx' required></label><div class='full'><button class='btn' type='submit'>Importar e validar</button></div></form></div>"""
    return page("Importacao XLSX", body)


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {"á":"a","à":"a","ã":"a","â":"a","é":"e","ê":"e","í":"i","ó":"o","ô":"o","õ":"o","ú":"u","ç":"c"}
    for old, new in replacements.items():
        text = text.replace(old, new)
    return "".join(ch if ch.isalnum() else "_" for ch in text).strip("_")


@app.post("/imports", response_class=HTMLResponse)
async def imports_create(file: UploadFile = File(...)) -> HTMLResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return imports(error="Selecione um arquivo XLSX valido.")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        return imports(error="Arquivo superior a 20 MB.")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except Exception as exc:
        return imports(error=f"Falha ao abrir planilha: {exc}")
    headers = [normalize_header(v) for v in header_values]
    aliases = {
        "company": ["empresa", "razao_social", "company"], "cnpj": ["cnpj"], "category": ["categoria", "classe_do_risco", "risco"],
        "product": ["produto", "produto_utilizado", "nome_comercial"], "substance": ["substancia", "substancia_quimica", "agente_quimico"],
        "cas": ["cas", "numero_cas"], "t24": ["codigo_tabela_24", "tabela_24", "codigo_esocial"], "evaluation": ["avaliacao", "tipo_avaliacao"],
        "intensity": ["intensidade", "concentracao", "intensidade_concentracao"], "unit": ["unidade", "unidade_medida"],
        "technique": ["tecnica", "tecnica_medicao", "metodologia"], "epi": ["epi"], "epc": ["epc"],
        "t27": ["codigo_tabela_27", "tabela_27"], "exam": ["exame", "exame_ocupacional"]
    }
    index: dict[str, int] = {}
    for key, names in aliases.items():
        for name in names:
            if name in headers:
                index[key] = headers.index(name)
                break
    if "company" not in index and "cnpj" not in index:
        return imports(error="Planilha precisa conter Empresa ou CNPJ.")
    imported = 0
    skipped = 0
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in values):
            continue
        def val(key: str) -> str:
            pos = index.get(key)
            return "" if pos is None or pos >= len(values) or values[pos] is None else str(values[pos]).strip()
        company_name = val("company") or "Empresa importada"
        cnpj_digits = "".join(ch for ch in val("cnpj") if ch.isdigit())
        if len(cnpj_digits) != 14:
            skipped += 1
            continue
        company = one("SELECT * FROM companies WHERE cnpj=?", (cnpj_digits,))
        if not company:
            with db() as conn:
                cur = conn.execute("INSERT INTO companies(legal_name,cnpj,status,created_at) VALUES(?,?,?,?)", (company_name, cnpj_digits, "ativa", utcnow()))
                company_id = cur.lastrowid
        else:
            company_id = company["id"]
        category = val("category").lower() or "quimico"
        substance = val("substance")
        table24 = val("t24")
        evaluation = val("evaluation").lower() or "qualitativa"
        if category.startswith("quim") and (not substance or not table24):
            skipped += 1
            continue
        if evaluation.startswith("quant") and (not val("intensity") or not val("unit") or not val("technique")):
            skipped += 1
            continue
        with db() as conn:
            conn.execute("""INSERT INTO risks(company_id,worker_id,category,product_name,substance,cas_number,table24_code,exposure_form,exposure_route,frequency,evaluation_type,intensity,unit,technique,epi,epc,table27_code,exam_name,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (company_id,None,category,val("product"),substance,val("cas"),table24,"","","",evaluation,val("intensity"),val("unit"),val("technique"),val("epi"),val("epc"),val("t27"),val("exam"),utcnow()))
        imported += 1
    audit("IMPORT", "xlsx", file.filename, f"importados={imported};ignorados={skipped}")
    return imports(message=f"Importacao concluida: {imported} riscos importados; {skipped} linhas ignoradas por ausencia de dados obrigatorios.")


@app.get("/certificate", response_class=HTMLResponse)
def certificate(message: str = "", error: str = "") -> HTMLResponse:
    notice = f"<div class='alert ok'>{message}</div>" if message else (f"<div class='alert error'>{esc(error)}</div>" if error else "")
    body = f"""{notice}<div class='card'><h3>Conferir certificado A1</h3><div class='alert'>O arquivo e a senha sao processados apenas em memoria nesta requisicao e nao sao gravados no banco.</div><form method='post' enctype='multipart/form-data'><label class='full'>Arquivo PFX/P12<input type='file' name='file' accept='.pfx,.p12' required></label><label>Senha do certificado<input type='password' name='password' required></label><div><button class='btn' type='submit'>Conferir metadados</button></div></form></div>"""
    return page("Certificado digital A1", body)


@app.post("/certificate", response_class=HTMLResponse)
async def certificate_check(file: UploadFile = File(...), password: str = Form(...)) -> HTMLResponse:
    if not file.filename or Path(file.filename).suffix.lower() not in {".pfx", ".p12"}:
        return certificate(error="Arquivo deve ser PFX ou P12.")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        return certificate(error="Certificado superior a 10 MB.")
    try:
        key, cert, extra = pkcs12.load_key_and_certificates(content, password.encode("utf-8"))
        if key is None or cert is None:
            raise ValueError("certificado ou chave privada ausente")
        subject = esc(cert.subject.rfc4514_string())
        issuer = esc(cert.issuer.rfc4514_string())
        serial = esc(cert.serial_number)
        valid_from = esc(getattr(cert, "not_valid_before_utc", cert.not_valid_before))
        valid_to = esc(getattr(cert, "not_valid_after_utc", cert.not_valid_after))
        detail = f"<b>Titular:</b> {subject}<br><b>Emissor:</b> {issuer}<br><b>Serial:</b> {serial}<br><b>Valido de:</b> {valid_from}<br><b>Valido ate:</b> {valid_to}<br><b>Cadeia adicional:</b> {len(extra or [])} certificado(s)"
        audit("CHECK", "certificate", "memory", "metadados conferidos; arquivo nao persistido")
        return certificate(message=detail)
    except Exception:
        return certificate(error="Nao foi possivel abrir o certificado. Confira o arquivo e a senha.")


@app.get("/readiness", response_class=HTMLResponse)
def readiness() -> HTMLResponse:
    company_count = one("SELECT COUNT(*) n FROM companies WHERE status='ativa'")["n"]
    worker_count = one("SELECT COUNT(*) n FROM workers")["n"]
    risk_count = one("SELECT COUNT(*) n FROM risks")["n"]
    event_count = one("SELECT COUNT(*) n FROM events WHERE status='validado_local'")["n"]
    items = [
        (company_count > 0, "Empresa ativa cadastrada"),
        (worker_count > 0, "Trabalhador cadastrado"),
        (risk_count > 0, "Base de riscos cadastrada"),
        (event_count > 0, "Evento local validado"),
        (False, "XSD oficial vigente instalado"),
        (False, "Assinatura XMLDSig validada"),
        (False, "Web Service de Producao Restrita configurado"),
        (False, "Procuracao e certificado autorizados"),
    ]
    trs = "".join(f"<tr><td>{'OK' if ok else 'PENDENTE'}</td><td>{esc(label)}</td></tr>" for ok, label in items)
    ready_local = all(ok for ok, _ in items[:4])
    body = f"""<div class='alert {'ok' if ready_local else ''}'>{'Pronto para validacao local controlada.' if ready_local else 'Ainda existem pendencias cadastrais locais.'}</div><table><thead><tr><th>Status</th><th>Verificacao</th></tr></thead><tbody>{trs}</tbody></table><div class='alert error'>A interface nao declara prontidao para transmissao oficial enquanto XSD, assinatura, mTLS, procuracao e Producao Restrita nao estiverem validados.</div>"""
    return page("Central de prontidao", body)


@app.get("/audit", response_class=HTMLResponse)
def audit_page() -> HTMLResponse:
    data = rows("SELECT * FROM audit ORDER BY id DESC LIMIT 300")
    trs = "".join(f"<tr><td>{esc(r['created_at'])}</td><td>{esc(r['action'])}</td><td>{esc(r['entity'])}</td><td>{esc(r['entity_id'])}</td><td>{esc(r['detail'])}</td></tr>" for r in data) or "<tr><td colspan='5'>Nenhum registro.</td></tr>"
    return page("Trilha de auditoria", f"<table><thead><tr><th>Data</th><th>Acao</th><th>Entidade</th><th>ID</th><th>Detalhe</th></tr></thead><tbody>{trs}</tbody></table>")


@app.exception_handler(404)
def not_found(_, __) -> HTMLResponse:
    return page("Pagina nao encontrada", "<div class='alert error'>A rota solicitada nao existe.</div>")
