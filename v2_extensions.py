from __future__ import annotations

import hashlib
import html
import json
import os
import re
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook
from pypdf import PdfReader

FORBIDDEN_SUFFIXES = {".pfx", ".p12", ".pem", ".key", ".cer", ".crt"}
REFERENCE_SUFFIXES = {".xlsx", ".xlsm", ".pdf", ".txt", ".md", ".png", ".jpg", ".jpeg", ".webp"}
APP_VERSION = "2.0.0-rc1"

_module_path = Path(__file__).resolve()
ROOT_DIR = Path(os.getenv("SAENG_V2_ROOT", str(_module_path.parent.parent if _module_path.parent.name == "app" else _module_path.parent)))
STORAGE_DIR = ROOT_DIR / "storage"
REFERENCE_DIR = ROOT_DIR / "imports" / "references"
DOCS_DIR = ROOT_DIR / "docs"
DB_PATH = STORAGE_DIR / "saeng_v2_extensions.db"

for directory in (STORAGE_DIR, REFERENCE_DIR, DOCS_DIR):
    directory.mkdir(parents=True, exist_ok=True)
    keep = directory / ".keep"
    if not any(directory.iterdir()):
        keep.write_text("Diretório operacional preservado pelo SAENG Software SST V2.\n", encoding="utf-8")

router = APIRouter(prefix="/v2", tags=["SAENG V2"])


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def _db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_database() -> None:
    with _db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS source_files(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL UNIQUE,
                relative_path TEXT NOT NULL,
                kind TEXT NOT NULL,
                sha256 TEXT NOT NULL,
                size_bytes INTEGER NOT NULL,
                metadata_json TEXT NOT NULL DEFAULT '{}',
                scanned_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS workbook_sheets(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_id INTEGER NOT NULL REFERENCES source_files(id) ON DELETE CASCADE,
                sheet_name TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                column_count INTEGER NOT NULL,
                headers_json TEXT NOT NULL DEFAULT '[]',
                UNIQUE(source_id, sheet_name)
            );
            CREATE TABLE IF NOT EXISTS authorizations(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT,
                cnpj TEXT,
                valid_from TEXT,
                valid_until TEXT,
                services_json TEXT NOT NULL DEFAULT '[]',
                source_filename TEXT NOT NULL,
                verification_status TEXT NOT NULL DEFAULT 'DOCUMENTO_IMPORTADO',
                notes TEXT,
                scanned_at TEXT NOT NULL,
                UNIQUE(source_filename, cnpj)
            );
            CREATE TABLE IF NOT EXISTS extension_audit(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                details TEXT,
                created_at TEXT NOT NULL
            );
            """
        )


init_database()


def _require_user(request: Request) -> str:
    user = request.session.get("user") if hasattr(request, "session") else None
    if not user:
        raise HTTPException(status_code=401)
    return str(user)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _safe_text(path: Path, limit: int = 4_000_000) -> str:
    if path.stat().st_size > limit:
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def inspect_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            first_rows = list(worksheet.iter_rows(min_row=1, max_row=min(12, worksheet.max_row or 1), values_only=True))
            header_values: list[str] = []
            for row in first_rows:
                values = [str(value).strip() if value is not None else "" for value in row]
                if sum(bool(value) for value in values) >= 2:
                    header_values = values
                    break
            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": int(worksheet.max_row or 0),
                    "columns": int(worksheet.max_column or 0),
                    "headers": header_values[:80],
                }
            )
    finally:
        workbook.close()
    return {"sheet_count": len(sheets), "sheets": sheets}


def _pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _normalize_cnpj(value: str | None) -> str | None:
    digits = re.sub(r"\D", "", value or "")
    return digits if len(digits) == 14 else None


def extract_authorization(path: Path) -> dict[str, Any] | None:
    try:
        text = _pdf_text(path)
    except Exception:
        return None
    upper = text.upper()
    if "PROCURAÇÃO" not in upper and "PROCURACAO" not in upper and "AUTORIZAÇÃO DE ACESSO" not in upper and "AUTORIZACAO DE ACESSO" not in upper:
        return None

    cnpjs = [_normalize_cnpj(match) for match in re.findall(r"\d{2}[.]?\d{3}[.]?\d{3}[/]?\d{4}[-]?\d{2}", text)]
    cnpjs = [value for value in cnpjs if value]
    grantor_cnpj = cnpjs[0] if cnpjs else None

    company_patterns = [
        r"Outorgante:\s*(?:A empresa\s+)?(.+?),\s*(?:também denominada|CNPJ)",
        r"representante de\s+(.+?),\s*inscrita no CNPJ",
    ]
    company_name = None
    for pattern in company_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            company_name = re.sub(r"\s+", " ", match.group(1)).strip(" ,.-")[:220]
            break

    date_strings = re.findall(r"\b\d{2}/\d{2}/\d{4}\b", text)
    valid_from = date_strings[0] if date_strings else None
    valid_until = date_strings[1] if len(date_strings) > 1 else None
    explicit_until = re.search(r"(?:até|ate)\s+(\d{2}/\d{2}/\d{4})", text, flags=re.IGNORECASE)
    if explicit_until:
        valid_until = explicit_until.group(1)

    services: list[str] = []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip(" -•\t")
        if "ESOCIAL" in line.upper() or "TODOS OS SERVIÇOS" in line.upper() or "TODOS OS SERVICOS" in line.upper():
            if 3 <= len(line) <= 300 and line not in services:
                services.append(line)

    return {
        "company_name": company_name,
        "cnpj": grantor_cnpj,
        "valid_from": valid_from,
        "valid_until": valid_until,
        "services": services[:30],
        "verification_status": "DOCUMENTO_IMPORTADO_NAO_CONSULTADO",
    }


def classify_file(path: Path) -> str:
    suffix = path.suffix.lower()
    name = path.name.lower()
    if suffix in {".xlsx", ".xlsm"}:
        if "aso" in name:
            return "PLANILHA_ASO"
        if "esocial" in name and "envio" in name:
            return "PLANILHA_ENVIOS_ESOCIAL"
        if "risco" in name:
            return "PLANILHA_RISCOS"
        if "finance" in name:
            return "PLANILHA_FINANCEIRA"
        if "agenda" in name:
            return "PLANILHA_AGENDA"
        if "sms" in name or "sst" in name:
            return "PLANILHA_CONTROLE_SST"
        return "PLANILHA"
    if suffix == ".pdf":
        return "DOCUMENTO_PDF"
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return "IDENTIDADE_VISUAL"
    if suffix in {".txt", ".md"}:
        return "REFERENCIA_TEXTO"
    return "REFERENCIA"


def scan_reference_folder() -> dict[str, Any]:
    results = {"processed": 0, "ignored": 0, "errors": []}
    with _db() as conn:
        for path in sorted(REFERENCE_DIR.rglob("*")):
            if not path.is_file() or path.name == ".keep":
                continue
            if path.suffix.lower() in FORBIDDEN_SUFFIXES or "senha" in path.name.lower():
                results["ignored"] += 1
                continue
            if path.suffix.lower() not in REFERENCE_SUFFIXES:
                results["ignored"] += 1
                continue
            try:
                metadata: dict[str, Any] = {}
                if path.suffix.lower() in {".xlsx", ".xlsm"}:
                    metadata = inspect_workbook(path)
                elif path.suffix.lower() in {".txt", ".md"}:
                    metadata = {"characters": len(_safe_text(path))}

                conn.execute(
                    """
                    INSERT INTO source_files(filename,relative_path,kind,sha256,size_bytes,metadata_json,scanned_at)
                    VALUES(?,?,?,?,?,?,?)
                    ON CONFLICT(filename) DO UPDATE SET
                        relative_path=excluded.relative_path,
                        kind=excluded.kind,
                        sha256=excluded.sha256,
                        size_bytes=excluded.size_bytes,
                        metadata_json=excluded.metadata_json,
                        scanned_at=excluded.scanned_at
                    """,
                    (
                        path.name,
                        str(path.relative_to(ROOT_DIR)),
                        classify_file(path),
                        _sha256(path),
                        path.stat().st_size,
                        json.dumps(metadata, ensure_ascii=False),
                        _now(),
                    ),
                )
                source_id = conn.execute("SELECT id FROM source_files WHERE filename=?", (path.name,)).fetchone()["id"]
                conn.execute("DELETE FROM workbook_sheets WHERE source_id=?", (source_id,))
                for sheet in metadata.get("sheets", []):
                    conn.execute(
                        "INSERT INTO workbook_sheets(source_id,sheet_name,row_count,column_count,headers_json) VALUES(?,?,?,?,?)",
                        (source_id, sheet["name"], sheet["rows"], sheet["columns"], json.dumps(sheet["headers"], ensure_ascii=False)),
                    )

                if path.suffix.lower() == ".pdf":
                    authorization = extract_authorization(path)
                    if authorization:
                        conn.execute(
                            """
                            INSERT INTO authorizations(company_name,cnpj,valid_from,valid_until,services_json,source_filename,verification_status,notes,scanned_at)
                            VALUES(?,?,?,?,?,?,?,?,?)
                            ON CONFLICT(source_filename,cnpj) DO UPDATE SET
                                company_name=excluded.company_name,
                                valid_from=excluded.valid_from,
                                valid_until=excluded.valid_until,
                                services_json=excluded.services_json,
                                verification_status=excluded.verification_status,
                                scanned_at=excluded.scanned_at
                            """,
                            (
                                authorization.get("company_name"),
                                authorization.get("cnpj"),
                                authorization.get("valid_from"),
                                authorization.get("valid_until"),
                                json.dumps(authorization.get("services", []), ensure_ascii=False),
                                path.name,
                                authorization["verification_status"],
                                "A validade definitiva deve ser confirmada no serviço oficial.",
                                _now(),
                            ),
                        )
                results["processed"] += 1
            except Exception as exc:
                results["errors"].append(f"{path.name}: {exc}")
        conn.execute(
            "INSERT INTO extension_audit(action,details,created_at) VALUES(?,?,?)",
            ("SCAN_REFERENCES", json.dumps(results, ensure_ascii=False), _now()),
        )
    return results


def audit_project_tree() -> dict[str, Any]:
    required_files = [
        "app/main.py",
        "app/config.py",
        "app/database.py",
        "app/models.py",
        "requirements.txt",
        "START_SAENG_SST.bat",
        "EXECUTAR_TESTES.bat",
    ]
    required_dirs = [
        "app",
        "app/templates",
        "app/static",
        "app/esocial",
        "schemas",
        "storage",
        "storage/documents",
        "storage/xml",
        "storage/reports",
        "imports/references",
        "docs",
        "tests",
    ]
    missing_files = [item for item in required_files if not (ROOT_DIR / item).exists()]
    missing_dirs = [item for item in required_dirs if not (ROOT_DIR / item).is_dir()]
    empty_dirs: list[str] = []
    forbidden: list[str] = []
    for path in ROOT_DIR.rglob("*"):
        if path.is_file() and path.suffix.lower() in FORBIDDEN_SUFFIXES:
            forbidden.append(str(path.relative_to(ROOT_DIR)))
        if path.is_dir() and path.name not in {".git", ".venv", "__pycache__", ".pytest_cache"}:
            try:
                if not any(path.iterdir()):
                    empty_dirs.append(str(path.relative_to(ROOT_DIR)))
            except OSError:
                pass
    return {
        "required_files_missing": missing_files,
        "required_directories_missing": missing_dirs,
        "empty_directories": empty_dirs,
        "forbidden_secret_files": forbidden,
        "ok": not missing_files and not missing_dirs and not forbidden,
    }


def _layout(title: str, content: str) -> HTMLResponse:
    nav = "".join(
        f'<a href="{url}">{label}</a>'
        for url, label in [
            ("/", "Sistema principal"),
            ("/v2", "Visão V2"),
            ("/v2/sources", "Fontes importadas"),
            ("/v2/authorizations", "Autorizações"),
            ("/v2/risk-engine", "Motor de riscos"),
            ("/v2/system-audit", "Auditoria estrutural"),
        ]
    )
    css = """
    :root{--navy:#061b33;--navy2:#0b2949;--gold:#d1a33b;--bg:#f3f6fa;--ink:#142239;--muted:#607087;--line:#dce4ed;--ok:#17734a;--warn:#926814}
    *{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font:14px/1.5 Inter,Segoe UI,Arial,sans-serif}
    header{background:linear-gradient(135deg,var(--navy),var(--navy2));color:#fff;padding:24px 5vw;border-bottom:3px solid var(--gold)}
    header h1{margin:0;font-size:28px}header p{margin:4px 0 0;color:#cbd7e4}nav{display:flex;flex-wrap:wrap;gap:8px;margin-top:16px}nav a{color:#fff;text-decoration:none;border:1px solid #ffffff2d;padding:8px 11px;border-radius:9px}nav a:hover{border-color:var(--gold);color:#f5d77e}
    main{max-width:1320px;margin:auto;padding:26px 5vw 50px}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px}.card{background:#fff;border:1px solid var(--line);border-radius:15px;padding:17px;box-shadow:0 8px 24px #17324e0c}.metric b{display:block;font-size:30px;color:var(--navy)}.metric span,.muted{color:var(--muted)}table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line);border-radius:14px;overflow:hidden}th,td{padding:10px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}th{background:#eaf0f6}.btn{display:inline-block;background:var(--navy);color:#fff;border:0;border-radius:9px;padding:10px 14px;font-weight:700;cursor:pointer;text-decoration:none}.btn.gold{background:var(--gold);color:#182235}.notice{padding:12px 14px;border-radius:10px;background:#fff6db;border:1px solid #e3c875;color:#6f5414;margin:14px 0}.ok{color:var(--ok);font-weight:700}.bad{color:#9b3030;font-weight:700}code{background:#edf1f5;padding:2px 5px;border-radius:5px}.section{margin-top:20px}.chips{display:flex;flex-wrap:wrap;gap:7px}.chip{background:#edf2f7;border-radius:999px;padding:5px 9px;font-size:12px}
    @media(max-width:900px){.grid{grid-template-columns:repeat(2,1fr)}}@media(max-width:560px){.grid{grid-template-columns:1fr}header,main{padding-left:18px;padding-right:18px}table{display:block;overflow:auto}}
    """
    return HTMLResponse(f"<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{_esc(title)} · SAENG SST V2</title><style>{css}</style></head><body><header><h1>{_esc(title)}</h1><p>SAENG Software SST V2 · extensão local de governança, referências e validação</p><nav>{nav}</nav></header><main>{content}</main></body></html>")


@router.get("", response_class=HTMLResponse)
def v2_dashboard(request: Request) -> HTMLResponse:
    _require_user(request)
    with _db() as conn:
        sources = conn.execute("SELECT COUNT(*) n FROM source_files").fetchone()["n"]
        sheets = conn.execute("SELECT COUNT(*) n FROM workbook_sheets").fetchone()["n"]
        authorizations = conn.execute("SELECT COUNT(*) n FROM authorizations").fetchone()["n"]
    audit = audit_project_tree()
    metrics = [("Fontes", sources), ("Abas XLSX", sheets), ("Autorizações", authorizations), ("Estrutura", "OK" if audit["ok"] else "Revisar")]
    cards = "".join(f"<div class='card metric'><b>{_esc(value)}</b><span>{_esc(label)}</span></div>" for label, value in metrics)
    content = f"""
    <div class='notice'><strong>Segurança:</strong> certificado PFX/P12 e senha não são importados, gravados, copiados para backup ou incluídos no ZIP. A validação da procuração continua dependente do serviço oficial.</div>
    <div class='grid'>{cards}</div>
    <div class='section card'><h2>Fluxo mestre</h2><p>Empresa → autorização documentada → estabelecimento/GHE/função → trabalhador → documento técnico → risco/agente/substância → exame → evento → XML/XSD → assinatura em memória → lote → protocolo → consulta → recibo individual.</p>
    <form method='post' action='/v2/scan'><button class='btn gold'>Ler e indexar referências locais</button></form></div>
    """
    return _layout("Centro de governança V2", content)


@router.post("/scan")
def scan(request: Request) -> RedirectResponse:
    _require_user(request)
    scan_reference_folder()
    return RedirectResponse("/v2/sources", status_code=303)


@router.get("/sources", response_class=HTMLResponse)
def sources_page(request: Request) -> HTMLResponse:
    _require_user(request)
    with _db() as conn:
        sources = conn.execute("SELECT * FROM source_files ORDER BY kind, filename").fetchall()
        sheets = conn.execute("SELECT * FROM workbook_sheets ORDER BY source_id, sheet_name").fetchall()
    sheet_map: dict[int, list[sqlite3.Row]] = {}
    for sheet in sheets:
        sheet_map.setdefault(sheet["source_id"], []).append(sheet)
    rows_html = ""
    for source in sources:
        details = ""
        if source["id"] in sheet_map:
            details = "<div class='chips'>" + "".join(f"<span class='chip'>{_esc(sheet['sheet_name'])}: {sheet['row_count']}×{sheet['column_count']}</span>" for sheet in sheet_map[source["id"]]) + "</div>"
        rows_html += f"<tr><td>{_esc(source['kind'])}</td><td>{_esc(source['filename'])}{details}</td><td>{source['size_bytes']:,}</td><td><code>{_esc(source['sha256'][:16])}…</code></td></tr>"
    rows_html = rows_html or "<tr><td colspan='4'>Nenhuma referência indexada. Execute a leitura na Visão V2.</td></tr>"
    return _layout("Fontes e planilhas importadas", f"<table><thead><tr><th>Tipo</th><th>Arquivo/abas</th><th>Bytes</th><th>SHA-256</th></tr></thead><tbody>{rows_html}</tbody></table>")


@router.get("/authorizations", response_class=HTMLResponse)
def authorizations_page(request: Request) -> HTMLResponse:
    _require_user(request)
    with _db() as conn:
        records = conn.execute("SELECT * FROM authorizations ORDER BY valid_until, company_name").fetchall()
    rows_html = ""
    today = date.today()
    for item in records:
        services = json.loads(item["services_json"] or "[]")
        status = item["verification_status"]
        try:
            if item["valid_until"]:
                parsed = datetime.strptime(item["valid_until"], "%d/%m/%Y").date()
                if parsed < today:
                    status = "DOCUMENTO_VENCIDO"
        except ValueError:
            pass
        rows_html += f"<tr><td>{_esc(item['company_name'] or 'Não identificado')}</td><td>{_esc(item['cnpj'] or '—')}</td><td>{_esc(item['valid_from'] or '—')} a {_esc(item['valid_until'] or '—')}</td><td>{'<br>'.join(_esc(value) for value in services) or '—'}</td><td>{_esc(status)}</td><td>{_esc(item['source_filename'])}</td></tr>"
    rows_html = rows_html or "<tr><td colspan='6'>Nenhum documento de autorização foi identificado.</td></tr>"
    notice = "<div class='notice'>O sistema registra evidências e escopo documental. Ele não concede permissões, não contorna GOV.BR/MFA e não consulta automaticamente a página privada do e-CAC. A confirmação definitiva deve ser feita pelo operador no serviço oficial.</div>"
    return _layout("Autorizações e procurações", notice + f"<table><thead><tr><th>Empresa</th><th>CNPJ</th><th>Vigência</th><th>Serviços identificados</th><th>Status interno</th><th>Fonte</th></tr></thead><tbody>{rows_html}</tbody></table>")


@router.get("/risk-engine", response_class=HTMLResponse)
def risk_engine_page(request: Request) -> HTMLResponse:
    _require_user(request)
    rules = [
        "Produto químico não pode ser tratado apenas pelo nome comercial.",
        "Substância, CAS quando disponível e código da Tabela 24 devem ser vinculados ao agente.",
        "Avaliação quantitativa exige resultado, unidade e técnica/metodologia.",
        "EPC, EPI, documento/CA e eficácia devem ser avaliados por agente.",
        "Exames devem ser vinculados ao PCMSO e ao código da Tabela 27.",
        "O código 09.01.001 não significa ausência de todo risco ocupacional; é uma declaração previdenciária específica.",
        "Risco do PGR, agente previdenciário, insalubridade e exame médico são conceitos relacionados, porém distintos.",
    ]
    content = "<div class='grid'>" + "".join(f"<div class='card'>{_esc(rule)}</div>" for rule in rules) + "</div>"
    content += "<div class='section card'><h2>Encadeamento obrigatório</h2><p>Empresa → estabelecimento → setor → GHE → função/CBO → atividade → produto → substância → agente/Tabela 24 → avaliação → EPC/EPI → exame/Tabela 27 → S-2240/S-2220.</p></div>"
    return _layout("Motor mestre de riscos ocupacionais", content)


@router.get("/system-audit", response_class=HTMLResponse)
def system_audit_page(request: Request) -> HTMLResponse:
    _require_user(request)
    result = audit_project_tree()
    def list_items(items: list[str]) -> str:
        return "<ul>" + "".join(f"<li>{_esc(item)}</li>" for item in items) + "</ul>" if items else "<p class='ok'>Nenhum item.</p>"
    content = f"""
    <div class='card'><h2>Status geral</h2><p class='{'ok' if result['ok'] else 'bad'}'>{'Estrutura aprovada nesta verificação.' if result['ok'] else 'Existem pendências estruturais.'}</p></div>
    <div class='grid section'>
      <div class='card'><h3>Arquivos obrigatórios ausentes</h3>{list_items(result['required_files_missing'])}</div>
      <div class='card'><h3>Pastas obrigatórias ausentes</h3>{list_items(result['required_directories_missing'])}</div>
      <div class='card'><h3>Pastas vazias</h3>{list_items(result['empty_directories'])}</div>
      <div class='card'><h3>Segredos proibidos na pasta</h3>{list_items(result['forbidden_secret_files'])}</div>
    </div>
    """
    return _layout("Auditoria pasta por pasta", content)
