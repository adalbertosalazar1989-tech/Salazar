from __future__ import annotations

import hashlib
import html
import io
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook
from sqlalchemy import Boolean, Date, DateTime, Float, ForeignKey, Integer, String, Text, UniqueConstraint, func
from sqlalchemy.orm import Mapped, Session, mapped_column

from .config import ROOT_DIR, settings
from .database import Base, engine, get_db
from .esocial.xml_builder import generate_event_id
from .models import AuditLog, Company, Event, Worker

router = APIRouter(prefix="/operacional", tags=["SAENG V2 Operacional"])
IMPORT_DIR = ROOT_DIR / "imports" / "operacional"
IMPORT_DIR.mkdir(parents=True, exist_ok=True)


def now_naive() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


class EstablishmentV2(Base):
    __tablename__ = "v2_establishments"
    __table_args__ = (UniqueConstraint("company_id", "registration_number", name="uq_v2_est_company_reg"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    name: Mapped[str] = mapped_column(String(255))
    registration_type: Mapped[str] = mapped_column(String(2), default="1")
    registration_number: Mapped[str] = mapped_column(String(30))
    address: Mapped[str | None] = mapped_column(String(400), nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class GHEV2(Base):
    __tablename__ = "v2_ghe"
    __table_args__ = (UniqueConstraint("company_id", "name", name="uq_v2_ghe_company_name"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    establishment_id: Mapped[int | None] = mapped_column(ForeignKey("v2_establishments.id"), nullable=True)
    name: Mapped[str] = mapped_column(String(180))
    sector: Mapped[str | None] = mapped_column(String(180), nullable=True)
    description: Mapped[str | None] = mapped_column(Text, nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class FunctionV2(Base):
    __tablename__ = "v2_functions"
    __table_args__ = (UniqueConstraint("company_id", "name", "cbo", name="uq_v2_function_company"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    ghe_id: Mapped[int | None] = mapped_column(ForeignKey("v2_ghe.id"), nullable=True)
    name: Mapped[str] = mapped_column(String(220))
    cbo: Mapped[str | None] = mapped_column(String(16), nullable=True)
    activity_description: Mapped[str | None] = mapped_column(Text, nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class ChemicalProductV2(Base):
    __tablename__ = "v2_chemical_products"
    __table_args__ = (UniqueConstraint("company_id", "commercial_name", name="uq_v2_product_company_name"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    commercial_name: Mapped[str] = mapped_column(String(255))
    manufacturer: Mapped[str | None] = mapped_column(String(255), nullable=True)
    sds_reference: Mapped[str | None] = mapped_column(String(500), nullable=True)
    notes: Mapped[str | None] = mapped_column(Text, nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class ChemicalSubstanceV2(Base):
    __tablename__ = "v2_chemical_substances"
    __table_args__ = (UniqueConstraint("product_id", "technical_name", "cas_number", name="uq_v2_substance_product"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    product_id: Mapped[int] = mapped_column(ForeignKey("v2_chemical_products.id"), index=True)
    technical_name: Mapped[str] = mapped_column(String(255))
    cas_number: Mapped[str | None] = mapped_column(String(40), nullable=True)
    table24_code: Mapped[str] = mapped_column(String(30), index=True)
    table24_description: Mapped[str] = mapped_column(String(500))
    physical_form: Mapped[str | None] = mapped_column(String(80), nullable=True)
    exposure_route: Mapped[str | None] = mapped_column(String(180), nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class ExamCatalogV2(Base):
    __tablename__ = "v2_exam_catalog"
    __table_args__ = (UniqueConstraint("table27_code", "name", name="uq_v2_exam_code_name"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    table27_code: Mapped[str] = mapped_column(String(30), index=True)
    name: Mapped[str] = mapped_column(String(255))
    category: Mapped[str | None] = mapped_column(String(120), nullable=True)
    periodicity_months: Mapped[int | None] = mapped_column(Integer, nullable=True)
    medical_basis: Mapped[str | None] = mapped_column(Text, nullable=True)
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class AuthorizationV2(Base):
    __tablename__ = "v2_authorizations"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    services_json: Mapped[str] = mapped_column(Text, default="[]")
    valid_from: Mapped[date | None] = mapped_column(Date, nullable=True)
    valid_until: Mapped[date | None] = mapped_column(Date, nullable=True)
    status: Mapped[str] = mapped_column(String(40), default="NAO_VERIFICADA")
    evidence_filename: Mapped[str | None] = mapped_column(String(500), nullable=True)
    official_check_url: Mapped[str] = mapped_column(String(500), default="https://servicos.receitafederal.gov.br/servico/autorizacoes/minhas-autorizacoes")
    verified_at: Mapped[datetime | None] = mapped_column(DateTime, nullable=True)
    notes: Mapped[str | None] = mapped_column(Text, nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class RiskExposureV2(Base):
    __tablename__ = "v2_risk_exposures"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int] = mapped_column(ForeignKey("companies.id"), index=True)
    worker_id: Mapped[int | None] = mapped_column(ForeignKey("workers.id"), nullable=True, index=True)
    establishment_id: Mapped[int | None] = mapped_column(ForeignKey("v2_establishments.id"), nullable=True)
    ghe_id: Mapped[int | None] = mapped_column(ForeignKey("v2_ghe.id"), nullable=True)
    function_id: Mapped[int | None] = mapped_column(ForeignKey("v2_functions.id"), nullable=True)
    product_id: Mapped[int | None] = mapped_column(ForeignKey("v2_chemical_products.id"), nullable=True)
    substance_id: Mapped[int | None] = mapped_column(ForeignKey("v2_chemical_substances.id"), nullable=True)
    risk_class: Mapped[str] = mapped_column(String(40), default="QUIMICO")
    table24_code: Mapped[str] = mapped_column(String(30), index=True)
    table24_description: Mapped[str] = mapped_column(String(500))
    source_generator: Mapped[str | None] = mapped_column(String(400), nullable=True)
    activities: Mapped[str] = mapped_column(Text)
    frequency: Mapped[str | None] = mapped_column(String(80), nullable=True)
    exposure_time: Mapped[str | None] = mapped_column(String(80), nullable=True)
    assessment_type: Mapped[str] = mapped_column(String(20), default="QUALITATIVA")
    intensity: Mapped[float | None] = mapped_column(Float, nullable=True)
    unit: Mapped[str | None] = mapped_column(String(40), nullable=True)
    methodology: Mapped[str | None] = mapped_column(String(500), nullable=True)
    tolerance_limit: Mapped[float | None] = mapped_column(Float, nullable=True)
    action_level: Mapped[float | None] = mapped_column(Float, nullable=True)
    epc: Mapped[str | None] = mapped_column(Text, nullable=True)
    epi: Mapped[str | None] = mapped_column(Text, nullable=True)
    ca_document: Mapped[str | None] = mapped_column(String(200), nullable=True)
    epc_effective: Mapped[bool | None] = mapped_column(Boolean, nullable=True)
    epi_effective: Mapped[bool | None] = mapped_column(Boolean, nullable=True)
    exam_id: Mapped[int | None] = mapped_column(ForeignKey("v2_exam_catalog.id"), nullable=True)
    start_date: Mapped[date] = mapped_column(Date)
    end_date: Mapped[date | None] = mapped_column(Date, nullable=True)
    responsible_cpf: Mapped[str | None] = mapped_column(String(11), nullable=True)
    responsible_registry: Mapped[str | None] = mapped_column(String(80), nullable=True)
    responsible_state: Mapped[str | None] = mapped_column(String(2), nullable=True)
    status: Mapped[str] = mapped_column(String(40), default="RASCUNHO")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive, onupdate=now_naive)


class ImportBatchV2(Base):
    __tablename__ = "v2_import_batches"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    filename: Mapped[str] = mapped_column(String(500))
    kind: Mapped[str] = mapped_column(String(60))
    sha256: Mapped[str] = mapped_column(String(64), index=True)
    total_rows: Mapped[int] = mapped_column(Integer, default=0)
    imported_rows: Mapped[int] = mapped_column(Integer, default=0)
    error_rows: Mapped[int] = mapped_column(Integer, default=0)
    status: Mapped[str] = mapped_column(String(40), default="PROCESSANDO")
    report_json: Mapped[str] = mapped_column(Text, default="{}")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)


class InconsistencyV2(Base):
    __tablename__ = "v2_inconsistencies"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    company_id: Mapped[int | None] = mapped_column(ForeignKey("companies.id"), nullable=True, index=True)
    worker_id: Mapped[int | None] = mapped_column(ForeignKey("workers.id"), nullable=True, index=True)
    entity_type: Mapped[str] = mapped_column(String(80))
    entity_id: Mapped[str | None] = mapped_column(String(80), nullable=True)
    severity: Mapped[str] = mapped_column(String(20), default="ALTA")
    code: Mapped[str] = mapped_column(String(80), index=True)
    field_name: Mapped[str | None] = mapped_column(String(120), nullable=True)
    current_value: Mapped[str | None] = mapped_column(Text, nullable=True)
    rule: Mapped[str] = mapped_column(Text)
    suggestion: Mapped[str] = mapped_column(Text)
    status: Mapped[str] = mapped_column(String(30), default="ABERTA")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive)
    resolved_at: Mapped[datetime | None] = mapped_column(DateTime, nullable=True)


class HomologationCaseV2(Base):
    __tablename__ = "v2_homologation_cases"
    __table_args__ = (UniqueConstraint("event_type", "scenario", name="uq_v2_homologation_case"),)
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    event_type: Mapped[str] = mapped_column(String(20))
    scenario: Mapped[str] = mapped_column(String(255))
    environment: Mapped[str] = mapped_column(String(30), default="PRODUCAO_RESTRITA")
    status: Mapped[str] = mapped_column(String(40), default="PENDENTE")
    protocol: Mapped[str | None] = mapped_column(String(128), nullable=True)
    receipt: Mapped[str | None] = mapped_column(String(128), nullable=True)
    evidence_path: Mapped[str | None] = mapped_column(String(500), nullable=True)
    notes: Mapped[str | None] = mapped_column(Text, nullable=True)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=now_naive, onupdate=now_naive)


Base.metadata.create_all(bind=engine)


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def require_user(request: Request) -> str:
    user = request.session.get("user") if hasattr(request, "session") else None
    if not user:
        raise HTTPException(status_code=401)
    return str(user)


def audit(db: Session, request: Request, action: str, entity: str, entity_id: str | None, details: str) -> None:
    db.add(AuditLog(actor=require_user(request), action=action, entity_type=entity, entity_id=entity_id, details=details))


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def bool_value(value: str | None) -> bool | None:
    upper = (value or "").strip().upper()
    if upper in {"S", "SIM", "1", "TRUE"}:
        return True
    if upper in {"N", "NAO", "NÃO", "0", "FALSE"}:
        return False
    return None


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9áàâãéêíóôõúç]+", "_", text)
    return text.strip("_")


def pick(row: dict[str, Any], *aliases: str) -> Any:
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and row[key] not in {None, ""}:
            return row[key]
    return None


def options(items: list[Any], label) -> str:
    parts: list[str] = []
    for item in items:
        parts.append("<option value='%s'>%s</option>" % (item.id, esc(label(item))))
    return "".join(parts)


def layout(title: str, body: str, request: Request) -> HTMLResponse:
    user = require_user(request)
    nav_items = [
        ("/", "Painel principal"), ("/operacional", "Central operacional"),
        ("/operacional/estrutura", "Estrutura/GHE"), ("/operacional/produtos", "Produtos e substâncias"),
        ("/operacional/exames", "Exames"), ("/operacional/riscos", "Motor de riscos"),
        ("/operacional/importacao", "Importação integral"), ("/operacional/autorizacoes", "Autorizações"),
        ("/operacional/inconsistencias", "Inconsistências"), ("/operacional/homologacao", "Homologação"),
    ]
    nav = "".join("<a href='%s'>%s</a>" % (url, label) for url, label in nav_items)
    css = """
    :root{--navy:#061a31;--navy2:#0a3157;--gold:#d3a536;--gold2:#f0ca67;--bg:#f2f5f9;--ink:#13243b;--muted:#607087;--line:#d8e1eb;--ok:#147248;--bad:#a23434}
    *{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font:14px/1.5 Segoe UI,Arial,sans-serif}header{background:linear-gradient(135deg,var(--navy),var(--navy2));color:#fff;padding:22px 4vw;border-bottom:3px solid var(--gold)}header h1{margin:0;font-size:30px}header p{margin:4px 0;color:#c9d7e7}nav{display:flex;gap:7px;flex-wrap:wrap;margin-top:15px}nav a{color:#fff;text-decoration:none;border:1px solid #ffffff30;border-radius:9px;padding:8px 10px}nav a:hover{border-color:var(--gold);color:var(--gold2)}main{max-width:1500px;margin:auto;padding:24px 4vw 60px}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px}.grid2{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}.card{background:#fff;border:1px solid var(--line);border-radius:16px;padding:18px;box-shadow:0 12px 32px #112b4510}.metric b{display:block;font-size:30px;color:var(--navy)}.metric span,.muted{color:var(--muted)}table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line)}th,td{padding:10px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}th{background:#eaf0f6}.tablewrap{overflow:auto;border-radius:14px}.section{margin-top:18px}form.gridform{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}label{font-weight:700;color:#33465f}input,select,textarea{width:100%;margin-top:5px;padding:10px;border:1px solid #cbd6e2;border-radius:9px;background:#fff;color:#13243b}textarea{min-height:90px}.span2{grid-column:span 2}.span4{grid-column:span 4}button,.btn{display:inline-block;border:0;border-radius:9px;padding:10px 14px;background:var(--navy);color:#fff;font-weight:800;text-decoration:none;cursor:pointer}.gold{background:linear-gradient(135deg,var(--gold2),var(--gold));color:#13243b}.small{padding:6px 9px;font-size:12px}.notice{background:#fff6d8;border:1px solid #e4ca74;color:#6c5317;border-radius:10px;padding:12px 14px;margin:12px 0}.ok{color:var(--ok);font-weight:800}.bad{color:var(--bad);font-weight:800}.badge{display:inline-block;border-radius:999px;background:#e9eef4;padding:4px 8px;font-size:11px;font-weight:800}.actions{display:flex;gap:7px;flex-wrap:wrap}.topline{display:flex;justify-content:space-between;gap:15px;align-items:center}.mono{font-family:Consolas,monospace;font-size:12px}@media(max-width:1000px){.grid{grid-template-columns:repeat(2,1fr)}form.gridform{grid-template-columns:repeat(2,1fr)}.span4{grid-column:span 2}}@media(max-width:620px){.grid,.grid2,form.gridform{grid-template-columns:1fr}.span2,.span4{grid-column:span 1}header,main{padding-left:16px;padding-right:16px}}
    """
    page = "<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>%s · SAENG Software SST</title><style>%s</style></head><body><header><div class='topline'><div><h1>%s</h1><p>SAENG Software SST V2 operacional · banco integrado · usuário %s</p></div><span class='badge'>%s / tpAmb %s</span></div><nav>%s</nav></header><main>%s</main></body></html>" % (esc(title), css, esc(title), esc(user), esc(settings.transmission_mode), esc(settings.esocial_environment), nav, body)
    return HTMLResponse(page)


def add_inconsistency(db: Session, code: str, rule: str, suggestion: str, company_id: int | None = None, worker_id: int | None = None, entity_type: str = "IMPORTACAO", entity_id: str | None = None, field_name: str | None = None, current_value: str | None = None) -> None:
    db.add(InconsistencyV2(company_id=company_id, worker_id=worker_id, entity_type=entity_type, entity_id=entity_id, code=code, field_name=field_name, current_value=current_value, rule=rule, suggestion=suggestion))


def validate_exposure(data: dict[str, Any]) -> list[tuple[str, str, str, str | None]]:
    errors: list[tuple[str, str, str, str | None]] = []
    code = str(data.get("table24_code") or "").strip()
    product = str(data.get("product_name") or "").strip()
    substance = str(data.get("substance_name") or "").strip()
    assessment = str(data.get("assessment_type") or "QUALITATIVA").upper()
    if not code:
        errors.append(("RISCO_T24_AUSENTE", "Todo agente destinado ao S-2240 exige código da Tabela 24.", "Informar o código oficial vigente.", "table24_code"))
    if product and not substance:
        errors.append(("QUIMICO_SEM_SUBSTANCIA", "Produto químico não pode ser tratado apenas pelo nome comercial.", "Cadastrar substância/composição e CAS quando disponível.", "substance_id"))
    if code == "09.01.001" and any(data.get(key) for key in ("epc", "epi", "ca_document", "intensity")):
        errors.append(("AUSENCIA_COM_CONTROLE", "O código 09.01.001 não deve coexistir com agente, avaliação, EPC ou EPI no mesmo registro.", "Revisar se realmente não existe agente nocivo previdenciário.", "table24_code"))
    if assessment == "QUANTITATIVA":
        for field, label in (("intensity", "intensidade/concentração"), ("unit", "unidade"), ("methodology", "técnica/metodologia")):
            if not data.get(field):
                errors.append(("QUANTITATIVO_INCOMPLETO", "Avaliação quantitativa exige %s." % label, "Preencher %s antes de validar." % label, field))
    return errors


@router.get("", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    metrics = [
        ("Estabelecimentos", db.query(EstablishmentV2).count()), ("GHE", db.query(GHEV2).count()),
        ("Produtos", db.query(ChemicalProductV2).count()), ("Substâncias", db.query(ChemicalSubstanceV2).count()),
        ("Exposições", db.query(RiskExposureV2).count()), ("Exames", db.query(ExamCatalogV2).count()),
        ("Autorizações", db.query(AuthorizationV2).count()), ("Pendências", db.query(InconsistencyV2).filter(InconsistencyV2.status == "ABERTA").count()),
    ]
    cards = "".join("<div class='card metric'><b>%s</b><span>%s</span></div>" % (value, esc(label)) for label, value in metrics)
    body = "<div class='notice'><b>Estado real:</b> esta central integra cadastros e gera rascunhos no banco principal. A aceitação oficial só existe após processamento e recibo individual do eSocial.</div><div class='grid'>%s</div>" % cards
    body += "<div class='grid2 section'><div class='card'><h2>Fluxo operacional</h2><p>Empresa → autorização → estabelecimento → GHE → função/CBO → produto → substância/CAS → Tabela 24 → avaliação → EPC/EPI → exame/Tabela 27 → S-2240/S-2220.</p></div><div class='card'><h2>Portais oficiais</h2><div class='actions'><a class='btn' target='_blank' rel='noopener' href='https://cav.receita.fazenda.gov.br/autenticacao/login'>Abrir e-CAC</a><a class='btn gold' target='_blank' rel='noopener' href='https://servicos.receitafederal.gov.br/servico/autorizacoes/minhas-autorizacoes'>Minhas autorizações</a></div><p class='muted'>Sem captura de senha GOV.BR, automação de MFA ou scraping de sessão privada.</p></div></div>"
    return layout("Central operacional integrada", body, request)


@router.get("/estrutura", response_class=HTMLResponse)
def structure_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    companies = db.query(Company).order_by(Company.razao_social).all()
    establishments = db.query(EstablishmentV2).order_by(EstablishmentV2.name).all()
    ghes = db.query(GHEV2).order_by(GHEV2.name).all()
    company_options = options(companies, lambda item: item.nome_fantasia or item.razao_social)
    est_options = options(establishments, lambda item: item.name)
    ghe_options = options(ghes, lambda item: item.name)
    body = """
    <div class='grid2'>
      <div class='card'><h2>Novo estabelecimento</h2><form class='gridform' method='post' action='/operacional/estrutura/estabelecimento'><label class='span2'>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label class='span2'>Nome<input name='name' required></label><label>Tipo inscrição<select name='registration_type'><option value='1'>CNPJ</option><option value='3'>CAEPF</option><option value='4'>CNO</option></select></label><label>Inscrição<input name='registration_number' required></label><label class='span2'>Endereço<input name='address'></label><button class='gold span2'>Salvar estabelecimento</button></form></div>
      <div class='card'><h2>Novo GHE</h2><form class='gridform' method='post' action='/operacional/estrutura/ghe'><label class='span2'>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label class='span2'>Estabelecimento<select name='establishment_id'><option value=''>Opcional</option>%s</select></label><label>Nome GHE<input name='name' required></label><label>Setor<input name='sector'></label><label class='span2'>Descrição<textarea name='description'></textarea></label><button class='gold span2'>Salvar GHE</button></form></div>
      <div class='card'><h2>Nova função</h2><form class='gridform' method='post' action='/operacional/estrutura/funcao'><label class='span2'>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label class='span2'>GHE<select name='ghe_id'><option value=''>Opcional</option>%s</select></label><label>Função<input name='name' required></label><label>CBO<input name='cbo'></label><label class='span2'>Atividades<textarea name='activity_description'></textarea></label><button class='gold span2'>Salvar função</button></form></div>
      <div class='card'><h2>Resumo estrutural</h2><p><b>%s</b> estabelecimentos · <b>%s</b> GHE · <b>%s</b> funções.</p></div>
    </div>""" % (company_options, company_options, est_options, company_options, ghe_options, len(establishments), len(ghes), db.query(FunctionV2).count())
    return layout("Estrutura, setores, GHE e funções", body, request)


@router.post("/estrutura/estabelecimento")
def create_establishment(request: Request, company_id: int = Form(...), name: str = Form(...), registration_type: str = Form("1"), registration_number: str = Form(...), address: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = EstablishmentV2(company_id=company_id, name=name.strip(), registration_type=registration_type, registration_number=re.sub(r"\W", "", registration_number.upper()), address=address.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_ESTABELECIMENTO", "EstablishmentV2", None, name); db.commit()
    return RedirectResponse("/operacional/estrutura", status_code=303)


@router.post("/estrutura/ghe")
def create_ghe(request: Request, company_id: int = Form(...), establishment_id: str = Form(""), name: str = Form(...), sector: str = Form(""), description: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = GHEV2(company_id=company_id, establishment_id=int(establishment_id) if establishment_id else None, name=name.strip(), sector=sector.strip() or None, description=description.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_GHE", "GHEV2", None, name); db.commit()
    return RedirectResponse("/operacional/estrutura", status_code=303)


@router.post("/estrutura/funcao")
def create_function(request: Request, company_id: int = Form(...), ghe_id: str = Form(""), name: str = Form(...), cbo: str = Form(""), activity_description: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = FunctionV2(company_id=company_id, ghe_id=int(ghe_id) if ghe_id else None, name=name.strip(), cbo=re.sub(r"\D", "", cbo) or None, activity_description=activity_description.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_FUNCAO", "FunctionV2", None, name); db.commit()
    return RedirectResponse("/operacional/estrutura", status_code=303)


@router.get("/produtos", response_class=HTMLResponse)
def products_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    companies = db.query(Company).order_by(Company.razao_social).all()
    products = db.query(ChemicalProductV2).order_by(ChemicalProductV2.commercial_name).all()
    substances = db.query(ChemicalSubstanceV2).order_by(ChemicalSubstanceV2.technical_name).all()
    rows: list[str] = []
    for item in substances:
        rows.append("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % (esc(item.technical_name), esc(item.cas_number or "—"), esc(item.table24_code), esc(item.table24_description), esc(item.physical_form or "—"), esc(item.exposure_route or "—")))
    rows_html = "".join(rows) or "<tr><td colspan='6'>Nenhuma substância cadastrada.</td></tr>"
    body = """
    <div class='grid2'>
      <div class='card'><h2>Produto comercial</h2><form class='gridform' method='post' action='/operacional/produtos/produto'><label class='span2'>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label class='span2'>Nome comercial<input name='commercial_name' required></label><label>Fabricante<input name='manufacturer'></label><label>FISPQ/FDS referência<input name='sds_reference'></label><label class='span2'>Observações<textarea name='notes'></textarea></label><button class='gold span2'>Salvar produto</button></form></div>
      <div class='card'><h2>Substância/composição</h2><form class='gridform' method='post' action='/operacional/produtos/substancia'><label class='span2'>Produto<select name='product_id' required><option value=''>Selecione</option>%s</select></label><label>Nome técnico<input name='technical_name' required></label><label>CAS<input name='cas_number'></label><label>Código Tabela 24<input name='table24_code' required></label><label class='span2'>Descrição Tabela 24<input name='table24_description' required></label><label>Forma física<input name='physical_form'></label><label>Via de exposição<input name='exposure_route'></label><button class='gold span2'>Salvar substância</button></form></div>
    </div><div class='section tablewrap'><table><thead><tr><th>Substância</th><th>CAS</th><th>Tabela 24</th><th>Descrição</th><th>Forma</th><th>Via</th></tr></thead><tbody>%s</tbody></table></div>""" % (options(companies, lambda item: item.nome_fantasia or item.razao_social), options(products, lambda item: item.commercial_name), rows_html)
    return layout("Produtos químicos e substâncias", body, request)


@router.post("/produtos/produto")
def create_product(request: Request, company_id: int = Form(...), commercial_name: str = Form(...), manufacturer: str = Form(""), sds_reference: str = Form(""), notes: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = ChemicalProductV2(company_id=company_id, commercial_name=commercial_name.strip(), manufacturer=manufacturer.strip() or None, sds_reference=sds_reference.strip() or None, notes=notes.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_PRODUTO", "ChemicalProductV2", None, commercial_name); db.commit()
    return RedirectResponse("/operacional/produtos", status_code=303)


@router.post("/produtos/substancia")
def create_substance(request: Request, product_id: int = Form(...), technical_name: str = Form(...), cas_number: str = Form(""), table24_code: str = Form(...), table24_description: str = Form(...), physical_form: str = Form(""), exposure_route: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = ChemicalSubstanceV2(product_id=product_id, technical_name=technical_name.strip(), cas_number=cas_number.strip() or None, table24_code=table24_code.strip(), table24_description=table24_description.strip(), physical_form=physical_form.strip() or None, exposure_route=exposure_route.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_SUBSTANCIA", "ChemicalSubstanceV2", None, technical_name); db.commit()
    return RedirectResponse("/operacional/produtos", status_code=303)


@router.get("/exames", response_class=HTMLResponse)
def exams_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    rows = []
    for item in db.query(ExamCatalogV2).order_by(ExamCatalogV2.table27_code, ExamCatalogV2.name).all():
        rows.append("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % (esc(item.table27_code), esc(item.name), esc(item.category or "—"), esc(item.periodicity_months or "—"), esc(item.medical_basis or "—")))
    rows_html = "".join(rows) or "<tr><td colspan='5'>Nenhum exame cadastrado.</td></tr>"
    body = "<div class='card'><h2>Novo exame/procedimento</h2><form class='gridform' method='post'><label>Código Tabela 27<input name='table27_code' required></label><label class='span2'>Nome<input name='name' required></label><label>Categoria<input name='category'></label><label>Periodicidade em meses<input name='periodicity_months' type='number' min='1'></label><label class='span2'>Base médica/PCMSO<textarea name='medical_basis'></textarea></label><button class='gold span2'>Salvar exame</button></form></div><div class='section tablewrap'><table><thead><tr><th>Tabela 27</th><th>Exame</th><th>Categoria</th><th>Periodicidade</th><th>Base</th></tr></thead><tbody>%s</tbody></table></div>" % rows_html
    return layout("Catálogo de exames ocupacionais", body, request)


@router.post("/exames")
def create_exam(request: Request, table27_code: str = Form(...), name: str = Form(...), category: str = Form(""), periodicity_months: str = Form(""), medical_basis: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    record = ExamCatalogV2(table27_code=table27_code.strip(), name=name.strip(), category=category.strip() or None, periodicity_months=int(periodicity_months) if periodicity_months else None, medical_basis=medical_basis.strip() or None)
    db.add(record); audit(db, request, "V2_CRIAR_EXAME", "ExamCatalogV2", None, "%s %s" % (table27_code, name)); db.commit()
    return RedirectResponse("/operacional/exames", status_code=303)


@router.get("/riscos", response_class=HTMLResponse)
def risks_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    companies = db.query(Company).order_by(Company.razao_social).all()
    workers = db.query(Worker).order_by(Worker.name).all()
    establishments = db.query(EstablishmentV2).order_by(EstablishmentV2.name).all()
    ghes = db.query(GHEV2).order_by(GHEV2.name).all()
    functions = db.query(FunctionV2).order_by(FunctionV2.name).all()
    products = db.query(ChemicalProductV2).order_by(ChemicalProductV2.commercial_name).all()
    substances = db.query(ChemicalSubstanceV2).order_by(ChemicalSubstanceV2.technical_name).all()
    exams = db.query(ExamCatalogV2).order_by(ExamCatalogV2.name).all()
    row_parts: list[str] = []
    for record in db.query(RiskExposureV2).order_by(RiskExposureV2.created_at.desc()).limit(500).all():
        company = db.get(Company, record.company_id)
        action = ""
        if record.worker_id:
            action = "<form method='post' action='/operacional/riscos/%s/gerar-s2240'><button class='small gold'>Gerar S-2240</button></form>" % record.id
        row_parts.append("<tr><td>#%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td><div class='actions'>%s<a class='btn small' href='/operacional/riscos/%s'>Detalhes</a></div></td></tr>" % (record.id, esc(company.nome_fantasia or company.razao_social), esc(record.risk_class), esc(record.table24_code), esc(record.table24_description), esc(record.assessment_type), esc(record.status), action, record.id))
    rows_html = "".join(row_parts) or "<tr><td colspan='8'>Nenhuma exposição cadastrada.</td></tr>"
    body = """
    <div class='notice'><b>Regra química:</b> nome comercial isolado não é suficiente. Informe substância, CAS quando disponível e código da Tabela 24. Avaliação quantitativa exige resultado, unidade e técnica.</div>
    <div class='card'><h2>Nova exposição</h2><form class='gridform' method='post'>
    <label>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label>Trabalhador<select name='worker_id'><option value=''>Aplicação por GHE/função</option>%s</select></label><label>Estabelecimento<select name='establishment_id'><option value=''>Opcional</option>%s</select></label><label>GHE<select name='ghe_id'><option value=''>Opcional</option>%s</select></label>
    <label>Função<select name='function_id'><option value=''>Opcional</option>%s</select></label><label>Produto<select name='product_id'><option value=''>Opcional</option>%s</select></label><label>Substância<select name='substance_id'><option value=''>Opcional</option>%s</select></label><label>Classe<select name='risk_class'><option>QUIMICO</option><option>FISICO</option><option>BIOLOGICO</option><option>ERGONOMICO</option><option>ACIDENTE</option><option>PSICOSSOCIAL</option></select></label>
    <label>Código Tabela 24<input name='table24_code' required></label><label class='span2'>Descrição Tabela 24<input name='table24_description' required></label><label>Data início<input name='start_date' type='date' required></label><label class='span2'>Atividades<textarea name='activities' required></textarea></label><label class='span2'>Fonte geradora<textarea name='source_generator'></textarea></label>
    <label>Frequência<input name='frequency'></label><label>Tempo de exposição<input name='exposure_time'></label><label>Tipo avaliação<select name='assessment_type'><option>QUALITATIVA</option><option>QUANTITATIVA</option></select></label><label>Intensidade<input name='intensity' type='number' step='any'></label><label>Unidade<input name='unit'></label><label class='span2'>Metodologia<input name='methodology'></label><label>Limite tolerância<input name='tolerance_limit' type='number' step='any'></label><label>Nível de ação<input name='action_level' type='number' step='any'></label>
    <label class='span2'>EPC<textarea name='epc'></textarea></label><label class='span2'>EPI<textarea name='epi'></textarea></label><label>CA/documento<input name='ca_document'></label><label>EPC eficaz?<select name='epc_effective'><option value=''>Não avaliado</option><option value='S'>Sim</option><option value='N'>Não</option></select></label><label>EPI eficaz?<select name='epi_effective'><option value=''>Não avaliado</option><option value='S'>Sim</option><option value='N'>Não</option></select></label><label>Exame<select name='exam_id'><option value=''>Opcional</option>%s</select></label>
    <label>CPF responsável<input name='responsible_cpf'></label><label>Registro profissional<input name='responsible_registry'></label><label>UF<input name='responsible_state' maxlength='2'></label><button class='gold span4'>Salvar e validar exposição</button></form></div>
    <div class='section tablewrap'><table><thead><tr><th>ID</th><th>Empresa</th><th>Classe</th><th>Tabela 24</th><th>Agente</th><th>Avaliação</th><th>Status</th><th>Ações</th></tr></thead><tbody>%s</tbody></table></div>""" % (
        options(companies, lambda item: item.nome_fantasia or item.razao_social), options(workers, lambda item: item.name), options(establishments, lambda item: item.name), options(ghes, lambda item: item.name), options(functions, lambda item: item.name), options(products, lambda item: item.commercial_name), options(substances, lambda item: "%s · %s" % (item.technical_name, item.table24_code)), options(exams, lambda item: "%s · %s" % (item.table27_code, item.name)), rows_html)
    return layout("Motor operacional de riscos", body, request)


@router.post("/riscos")
def create_risk(request: Request, company_id: int = Form(...), worker_id: str = Form(""), establishment_id: str = Form(""), ghe_id: str = Form(""), function_id: str = Form(""), product_id: str = Form(""), substance_id: str = Form(""), risk_class: str = Form("QUIMICO"), table24_code: str = Form(...), table24_description: str = Form(...), source_generator: str = Form(""), activities: str = Form(...), frequency: str = Form(""), exposure_time: str = Form(""), assessment_type: str = Form("QUALITATIVA"), intensity: str = Form(""), unit: str = Form(""), methodology: str = Form(""), tolerance_limit: str = Form(""), action_level: str = Form(""), epc: str = Form(""), epi: str = Form(""), ca_document: str = Form(""), epc_effective: str = Form(""), epi_effective: str = Form(""), exam_id: str = Form(""), start_date: str = Form(...), responsible_cpf: str = Form(""), responsible_registry: str = Form(""), responsible_state: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    worker = db.get(Worker, int(worker_id)) if worker_id else None
    if worker and worker.company_id != company_id:
        raise HTTPException(400, "O trabalhador não pertence à empresa selecionada.")
    product = db.get(ChemicalProductV2, int(product_id)) if product_id else None
    substance = db.get(ChemicalSubstanceV2, int(substance_id)) if substance_id else None
    validation_data = {"table24_code": table24_code.strip(), "product_name": product.commercial_name if product else None, "substance_name": substance.technical_name if substance else None, "assessment_type": assessment_type, "intensity": intensity, "unit": unit, "methodology": methodology, "epc": epc, "epi": epi, "ca_document": ca_document}
    errors = validate_exposure(validation_data)
    record = RiskExposureV2(company_id=company_id, worker_id=worker.id if worker else None, establishment_id=int(establishment_id) if establishment_id else None, ghe_id=int(ghe_id) if ghe_id else None, function_id=int(function_id) if function_id else None, product_id=int(product_id) if product_id else None, substance_id=int(substance_id) if substance_id else None, risk_class=risk_class, table24_code=table24_code.strip(), table24_description=table24_description.strip(), source_generator=source_generator.strip() or None, activities=activities.strip(), frequency=frequency.strip() or None, exposure_time=exposure_time.strip() or None, assessment_type=assessment_type, intensity=float(intensity) if intensity else None, unit=unit.strip() or None, methodology=methodology.strip() or None, tolerance_limit=float(tolerance_limit) if tolerance_limit else None, action_level=float(action_level) if action_level else None, epc=epc.strip() or None, epi=epi.strip() or None, ca_document=ca_document.strip() or None, epc_effective=bool_value(epc_effective), epi_effective=bool_value(epi_effective), exam_id=int(exam_id) if exam_id else None, start_date=parse_date(start_date), responsible_cpf=re.sub(r"\D", "", responsible_cpf) or None, responsible_registry=responsible_registry.strip() or None, responsible_state=responsible_state.upper().strip() or None, status="INVALIDO" if errors else "VALIDADO")
    db.add(record); db.flush()
    for code, rule, suggestion, field in errors:
        add_inconsistency(db, code, rule, suggestion, company_id=company_id, worker_id=record.worker_id, entity_type="RiskExposureV2", entity_id=str(record.id), field_name=field, current_value=str(validation_data.get(field) or ""))
    audit(db, request, "V2_CRIAR_EXPOSICAO", "RiskExposureV2", str(record.id), record.status); db.commit()
    return RedirectResponse("/operacional/riscos", status_code=303)


@router.get("/riscos/{risk_id}", response_class=HTMLResponse)
def risk_detail(request: Request, risk_id: int, db: Session = Depends(get_db)):
    require_user(request)
    record = db.get(RiskExposureV2, risk_id)
    if not record:
        raise HTTPException(404)
    worker = db.get(Worker, record.worker_id) if record.worker_id else None
    fields = [("Empresa", db.get(Company, record.company_id).razao_social), ("Trabalhador", worker.name if worker else "Aplicação por grupo"), ("Código T24", record.table24_code), ("Agente", record.table24_description), ("Atividades", record.activities), ("Avaliação", record.assessment_type), ("Intensidade", record.intensity), ("Unidade", record.unit), ("Metodologia", record.methodology), ("EPC", record.epc), ("EPI", record.epi), ("CA/documento", record.ca_document), ("Status", record.status)]
    rows = "".join("<tr><th>%s</th><td>%s</td></tr>" % (esc(key), esc(value or "—")) for key, value in fields)
    return layout("Exposição #%s" % risk_id, "<div class='card'><div class='tablewrap'><table><tbody>%s</tbody></table></div></div>" % rows, request)


@router.post("/riscos/{risk_id}/gerar-s2240")
def generate_s2240(request: Request, risk_id: int, db: Session = Depends(get_db)):
    require_user(request)
    record = db.get(RiskExposureV2, risk_id)
    if not record or not record.worker_id:
        raise HTTPException(400, "A exposição deve estar vinculada a um trabalhador.")
    if record.status != "VALIDADO":
        raise HTTPException(400, "Corrija as inconsistências antes de gerar o evento.")
    if not record.responsible_cpf or len(record.responsible_cpf) != 11:
        raise HTTPException(400, "Informe o CPF do responsável pelo registro ambiental.")
    company = db.get(Company, record.company_id)
    worker = db.get(Worker, record.worker_id)
    establishment = db.get(EstablishmentV2, record.establishment_id) if record.establishment_id else None
    ghe = db.get(GHEV2, record.ghe_id) if record.ghe_id else None
    payload = {
        "cnpj": company.cnpj, "nr_insc_empregador": company.cnpj[:8], "cpf_trab": worker.cpf, "matricula": worker.matricula,
        "cod_categ": worker.categoria, "tp_amb": settings.esocial_environment, "dt_ini_condicao": record.start_date.isoformat(),
        "dt_fim_condicao": record.end_date.isoformat() if record.end_date else None, "local_amb": "1",
        "dsc_setor": ghe.sector if ghe and ghe.sector else worker.setor or "Setor não informado",
        "tp_insc_amb": establishment.registration_type if establishment else "1",
        "nr_insc_amb": establishment.registration_number if establishment else company.establishment_cnpj or company.cnpj,
        "establishment_cnpj": company.establishment_cnpj or company.cnpj, "dsc_ativ_des": record.activities,
        "cod_ag_noc": record.table24_code, "dsc_ag_noc": record.table24_description,
        "tp_aval": "1" if record.assessment_type == "QUANTITATIVA" else "2", "int_conc": str(record.intensity or "0"),
        "lim_tol": str(record.tolerance_limit) if record.tolerance_limit is not None else None, "un_med": record.unit or "4",
        "tec_medicao": record.methodology or "Avaliação conforme documentação técnica",
        "utiliz_epc": "2" if record.epc else "1", "efic_epc": "S" if record.epc_effective else "N",
        "utiliz_epi": "2" if record.epi else "1", "efic_epi": "S" if record.epi_effective else "N", "ca_epi": record.ca_document,
        "cpf_resp": record.responsible_cpf, "ide_oc": "1" if record.responsible_registry else None,
        "nr_oc": record.responsible_registry, "uf_oc": record.responsible_state or "MG", "obs_compl": "Gerado pelo motor V2 a partir da exposição #%s." % record.id,
    }
    event = Event(company_id=company.id, worker_id=worker.id, event_type="S-2240", event_id=generate_event_id(company.cnpj, db.query(Event).count() % 99999 + 1), event_action="INCLUSAO", status="RASCUNHO", payload_json=json.dumps(payload, ensure_ascii=False))
    db.add(event); db.flush(); audit(db, request, "V2_GERAR_S2240", "Event", str(event.id), "exposicao=%s" % record.id); db.commit()
    return RedirectResponse("/events/%s" % event.id, status_code=303)


@router.get("/autorizacoes", response_class=HTMLResponse)
def authorizations_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    companies = db.query(Company).order_by(Company.razao_social).all()
    row_parts: list[str] = []
    for record in db.query(AuthorizationV2).order_by(AuthorizationV2.valid_until, AuthorizationV2.created_at.desc()).all():
        company = db.get(Company, record.company_id)
        services = ", ".join(json.loads(record.services_json or "[]"))
        row_parts.append("<tr><td>%s</td><td>%s</td><td>%s a %s</td><td><span class='badge'>%s</span></td><td>%s</td><td><a class='btn small' target='_blank' rel='noopener' href='%s'>Conferir oficial</a></td></tr>" % (esc(company.nome_fantasia or company.razao_social), esc(services), esc(record.valid_from or "—"), esc(record.valid_until or "—"), esc(record.status), esc(record.evidence_filename or "—"), esc(record.official_check_url)))
    rows_html = "".join(row_parts) or "<tr><td colspan='6'>Nenhuma autorização registrada.</td></tr>"
    body = """
    <div class='notice'>Sincronização permitida: importação de evidência e conferência assistida. O sistema não acessa conta GOV.BR, não guarda senha e não automatiza MFA.</div>
    <div class='card'><h2>Registrar autorização/procuração</h2><form class='gridform' method='post'><label class='span2'>Empresa<select name='company_id' required><option value=''>Selecione</option>%s</select></label><label class='span2'>Serviços, separados por ponto e vírgula<input name='services' placeholder='eSocial - Grupo SST; Download'></label><label>Válida de<input name='valid_from' type='date'></label><label>Válida até<input name='valid_until' type='date'></label><label>Status<select name='status'><option>NAO_VERIFICADA</option><option>VALIDA</option><option>PROXIMA_VENCIMENTO</option><option>VENCIDA</option></select></label><label>Arquivo/evidência<input name='evidence_filename'></label><label class='span4'>Observações<textarea name='notes'></textarea></label><button class='gold span4'>Salvar autorização</button></form></div>
    <div class='section tablewrap'><table><thead><tr><th>Empresa</th><th>Serviços</th><th>Vigência</th><th>Status</th><th>Evidência</th><th>Oficial</th></tr></thead><tbody>%s</tbody></table></div>""" % (options(companies, lambda item: item.nome_fantasia or item.razao_social), rows_html)
    return layout("Autorizações e procurações", body, request)


@router.post("/autorizacoes")
def create_authorization(request: Request, company_id: int = Form(...), services: str = Form(""), valid_from: str = Form(""), valid_until: str = Form(""), status: str = Form("NAO_VERIFICADA"), evidence_filename: str = Form(""), notes: str = Form(""), db: Session = Depends(get_db)):
    require_user(request)
    service_list = [item.strip() for item in services.split(";") if item.strip()]
    record = AuthorizationV2(company_id=company_id, services_json=json.dumps(service_list, ensure_ascii=False), valid_from=parse_date(valid_from), valid_until=parse_date(valid_until), status=status, evidence_filename=evidence_filename.strip() or None, verified_at=now_naive() if status == "VALIDA" else None, notes=notes.strip() or None)
    db.add(record); db.flush(); company = db.get(Company, company_id); company.authorization_status = status; company.authorization_valid_until = record.valid_until; audit(db, request, "V2_REGISTRAR_AUTORIZACAO", "AuthorizationV2", str(record.id), status); db.commit()
    return RedirectResponse("/operacional/autorizacoes", status_code=303)


def get_or_create_product(db: Session, company_id: int, name: str) -> ChemicalProductV2:
    record = db.query(ChemicalProductV2).filter(ChemicalProductV2.company_id == company_id, func.lower(ChemicalProductV2.commercial_name) == name.lower()).first()
    if not record:
        record = ChemicalProductV2(company_id=company_id, commercial_name=name); db.add(record); db.flush()
    return record


def get_or_create_substance(db: Session, product_id: int, name: str, cas: str | None, code: str, description: str) -> ChemicalSubstanceV2:
    record = db.query(ChemicalSubstanceV2).filter(ChemicalSubstanceV2.product_id == product_id, func.lower(ChemicalSubstanceV2.technical_name) == name.lower()).first()
    if not record:
        record = ChemicalSubstanceV2(product_id=product_id, technical_name=name, cas_number=cas or None, table24_code=code, table24_description=description); db.add(record); db.flush()
    return record


@router.get("/importacao", response_class=HTMLResponse)
def import_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    row_parts = []
    for batch in db.query(ImportBatchV2).order_by(ImportBatchV2.created_at.desc()).limit(100).all():
        row_parts.append("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % (esc(batch.created_at.strftime("%d/%m/%Y %H:%M")), esc(batch.filename), esc(batch.kind), batch.total_rows, batch.imported_rows, batch.error_rows, esc(batch.status)))
    rows_html = "".join(row_parts) or "<tr><td colspan='7'>Nenhuma importação executada.</td></tr>"
    body = """
    <div class='notice'>Importação integral: toda linha é importada ou registrada como inconsistência. Nenhuma linha é descartada silenciosamente.</div><div class='card'><h2>Importar XLSX</h2><form class='gridform' method='post' enctype='multipart/form-data'><label class='span2'>Tipo<select name='kind'><option value='RISCOS'>Planilha Mestra de Riscos</option><option value='ASO'>Controle de ASO</option></select></label><label class='span2'>Arquivo<input type='file' name='file' accept='.xlsx,.xlsm' required></label><button class='gold span4'>Processar integralmente</button></form></div><div class='section tablewrap'><table><thead><tr><th>Data</th><th>Arquivo</th><th>Tipo</th><th>Linhas</th><th>Importadas</th><th>Erros</th><th>Status</th></tr></thead><tbody>%s</tbody></table></div>""" % rows_html
    return layout("Importação integral de planilhas", body, request)


@router.post("/importacao")
async def import_xlsx(request: Request, kind: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    require_user(request)
    filename = re.sub(r"[^A-Za-z0-9._ -]", "_", Path(file.filename or "importacao.xlsx").name)[:180]
    if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(400, "Somente XLSX/XLSM.")
    data = await file.read(); (IMPORT_DIR / filename).write_bytes(data)
    batch = ImportBatchV2(filename=filename, kind=kind, sha256=hashlib.sha256(data).hexdigest()); db.add(batch); db.flush()
    total = imported = errors = 0; report: list[dict[str, Any]] = []
    workbook = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True); headers = next(iterator, None)
            if not headers:
                continue
            keys = [normalize_header(value) for value in headers]
            for row_number, values in enumerate(iterator, start=2):
                if not any(value not in {None, ""} for value in values):
                    continue
                total += 1; row = {keys[index]: values[index] if index < len(values) else None for index in range(len(keys)) if keys[index]}
                try:
                    if kind == "RISCOS":
                        cnpj = re.sub(r"\D", "", str(pick(row, "CNPJ", "CNPJ Empresa", "Empresa CNPJ") or "")); company = db.query(Company).filter(Company.cnpj == cnpj).first() if cnpj else None
                        if not company:
                            raise ValueError("Empresa/CNPJ não localizado no cadastro principal")
                        product_name = str(pick(row, "Produto", "Produto químico", "Nome comercial") or "").strip(); substance_name = str(pick(row, "Substância", "Agente químico", "Composição") or "").strip(); code24 = str(pick(row, "Código Tabela 24", "Tabela 24", "Cod Ag Noc") or "").strip(); description = str(pick(row, "Descrição Tabela 24", "Agente", "Risco") or "").strip(); activities = str(pick(row, "Atividade", "Atividades", "Descrição da atividade") or "").strip()
                        if not code24 or not activities:
                            raise ValueError("Código Tabela 24 e atividade são obrigatórios")
                        product = get_or_create_product(db, company.id, product_name) if product_name else None
                        substance = get_or_create_substance(db, product.id, substance_name, str(pick(row, "CAS") or "").strip() or None, code24, description or substance_name) if product and substance_name else None
                        exposure = RiskExposureV2(company_id=company.id, product_id=product.id if product else None, substance_id=substance.id if substance else None, risk_class=str(pick(row, "Classe do risco", "Tipo de risco") or "QUIMICO").upper(), table24_code=code24, table24_description=description or substance_name or "Agente não descrito", source_generator=str(pick(row, "Fonte geradora") or "").strip() or None, activities=activities, frequency=str(pick(row, "Frequência", "Frequencia") or "").strip() or None, exposure_time=str(pick(row, "Tempo de exposição", "Tempo exposicao") or "").strip() or None, assessment_type=str(pick(row, "Avaliação", "Tipo avaliação", "Avaliacao") or "QUALITATIVA").upper(), intensity=float(pick(row, "Intensidade", "Concentração", "Concentracao")) if pick(row, "Intensidade", "Concentração", "Concentracao") not in {None, ""} else None, unit=str(pick(row, "Unidade") or "").strip() or None, methodology=str(pick(row, "Técnica", "Metodologia", "Tecnica") or "").strip() or None, epc=str(pick(row, "EPC") or "").strip() or None, epi=str(pick(row, "EPI") or "").strip() or None, ca_document=str(pick(row, "CA", "Documento avaliação", "Documento avaliacao") or "").strip() or None, start_date=date.today(), status="RASCUNHO")
                        validation = validate_exposure({"table24_code": code24, "product_name": product_name, "substance_name": substance_name, "assessment_type": exposure.assessment_type, "intensity": exposure.intensity, "unit": exposure.unit, "methodology": exposure.methodology, "epc": exposure.epc, "epi": exposure.epi, "ca_document": exposure.ca_document}); exposure.status = "INVALIDO" if validation else "VALIDADO"; db.add(exposure); db.flush()
                        for code, rule, suggestion, field in validation:
                            add_inconsistency(db, code, rule, suggestion, company_id=company.id, entity_type="RiskExposureV2", entity_id=str(exposure.id), field_name=field)
                    else:
                        code27 = str(pick(row, "Código Tabela 27", "Tabela 27", "Código exame") or "").strip(); exam_name = str(pick(row, "Exame", "Procedimento", "Nome do exame") or "").strip()
                        if not code27 or not exam_name:
                            raise ValueError("Código Tabela 27 e nome do exame são obrigatórios")
                        exists = db.query(ExamCatalogV2).filter(ExamCatalogV2.table27_code == code27, func.lower(ExamCatalogV2.name) == exam_name.lower()).first()
                        if not exists:
                            db.add(ExamCatalogV2(table27_code=code27, name=exam_name, category=str(pick(row, "Categoria") or "").strip() or None, medical_basis=str(pick(row, "Base médica", "PCMSO", "Base medica") or "").strip() or None))
                    imported += 1
                except Exception as exc:
                    errors += 1; report.append({"sheet": worksheet.title, "row": row_number, "error": str(exc)}); add_inconsistency(db, "IMPORT_ROW_ERROR", "Toda linha deve ser importada ou registrada como pendência.", "Corrigir a linha e reprocessar.", entity_type="ImportBatchV2", entity_id=str(batch.id), current_value="%s!%s: %s" % (worksheet.title, row_number, exc))
        batch.total_rows = total; batch.imported_rows = imported; batch.error_rows = errors; batch.status = "CONCLUIDO_COM_PENDENCIAS" if errors else "CONCLUIDO"; batch.report_json = json.dumps(report, ensure_ascii=False); audit(db, request, "V2_IMPORTAR_PLANILHA", "ImportBatchV2", str(batch.id), "%s: %s/%s" % (kind, imported, total)); db.commit()
    finally:
        workbook.close()
    return RedirectResponse("/operacional/importacao", status_code=303)


@router.get("/inconsistencias", response_class=HTMLResponse)
def inconsistencies_page(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    rows = []
    for item in db.query(InconsistencyV2).order_by(InconsistencyV2.status, InconsistencyV2.created_at.desc()).limit(1000).all():
        action = ""
        if item.status == "ABERTA":
            action = "<form method='post' action='/operacional/inconsistencias/%s/resolver'><button class='small gold'>Resolver</button></form>" % item.id
        rows.append("<tr><td>#%s</td><td>%s</td><td>%s</td><td>%s %s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % (item.id, esc(item.severity), esc(item.code), esc(item.entity_type), esc(item.entity_id or ""), esc(item.field_name or "—"), esc(item.rule), esc(item.suggestion), esc(item.status), action))
    rows_html = "".join(rows) or "<tr><td colspan='9'>Nenhuma inconsistência.</td></tr>"
    body = "<div class='tablewrap'><table><thead><tr><th>ID</th><th>Gravidade</th><th>Código</th><th>Origem</th><th>Campo</th><th>Regra</th><th>Sugestão</th><th>Status</th><th>Ação</th></tr></thead><tbody>%s</tbody></table></div>" % rows_html
    return layout("Central de inconsistências", body, request)


@router.post("/inconsistencias/{item_id}/resolver")
def resolve_inconsistency(request: Request, item_id: int, db: Session = Depends(get_db)):
    require_user(request); item = db.get(InconsistencyV2, item_id)
    if not item:
        raise HTTPException(404)
    item.status = "RESOLVIDA"; item.resolved_at = now_naive(); audit(db, request, "V2_RESOLVER_INCONSISTENCIA", "InconsistencyV2", str(item.id), item.code); db.commit()
    return RedirectResponse("/operacional/inconsistencias", status_code=303)


HOMOLOGATION_SCENARIOS = [
    ("S-2210", "Inclusão válida"), ("S-2210", "Retificação"), ("S-2210", "Exclusão S-3000"),
    ("S-2220", "Inclusão ASO com múltiplos exames"), ("S-2220", "Retificação"), ("S-2220", "Exclusão S-3000"),
    ("S-2240", "Agente químico quantitativo"), ("S-2240", "Múltiplos agentes"), ("S-2240", "Código 09.01.001"),
    ("S-2240", "Retificação"), ("S-2240", "Exclusão S-3000"), ("LOTE", "Lote com até 50 eventos"),
    ("RETORNO", "Rejeição controlada"), ("RETORNO", "Advertência"), ("RETORNO", "Protocolo e recibo distintos"),
]


@router.get("/homologacao", response_class=HTMLResponse)
def homologation_page(request: Request, db: Session = Depends(get_db)):
    require_user(request); items = db.query(HomologationCaseV2).order_by(HomologationCaseV2.event_type, HomologationCaseV2.scenario).all(); rows = []
    for item in items:
        form = "<form class='actions' method='post' action='/operacional/homologacao/%s'><select name='status'><option>PENDENTE</option><option>EM_TESTE</option><option>APROVADO_RESTRITO</option><option>REJEITADO</option><option>APROVADO_PRODUCAO</option></select><input name='protocol' placeholder='protocolo'><input name='receipt' placeholder='recibo'><button class='small gold'>Atualizar</button></form>" % item.id
        rows.append("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td class='mono'>%s</td><td class='mono'>%s</td><td>%s</td></tr>" % (esc(item.event_type), esc(item.scenario), esc(item.environment), esc(item.status), esc(item.protocol or "—"), esc(item.receipt or "—"), form))
    rows_html = "".join(rows) or "<tr><td colspan='7'>Checklist ainda não criado.</td></tr>"; approved = sum(1 for item in items if item.status in {"APROVADO_RESTRITO", "APROVADO_PRODUCAO"})
    body = "<div class='notice'><b>Homologação oficial não pode ser fabricada localmente.</b> Exige certificado válido, autorização vigente, Produção Restrita e retorno real do Ambiente Nacional.</div><div class='grid'><div class='card metric'><b>%s</b><span>cenários</span></div><div class='card metric'><b>%s</b><span>aprovados</span></div><div class='card metric'><b>%s</b><span>pendentes</span></div><div class='card'><form method='post' action='/operacional/homologacao/inicializar'><button class='gold'>Criar checklist oficial</button></form></div></div><div class='section tablewrap'><table><thead><tr><th>Evento</th><th>Cenário</th><th>Ambiente</th><th>Status</th><th>Protocolo</th><th>Recibo</th><th>Atualizar</th></tr></thead><tbody>%s</tbody></table></div>" % (len(items), approved, len(items) - approved, rows_html)
    return layout("Homologação controlada", body, request)


@router.post("/homologacao/inicializar")
def seed_homologation(request: Request, db: Session = Depends(get_db)):
    require_user(request)
    for event_type, scenario in HOMOLOGATION_SCENARIOS:
        exists = db.query(HomologationCaseV2).filter(HomologationCaseV2.event_type == event_type, HomologationCaseV2.scenario == scenario).first()
        if not exists:
            db.add(HomologationCaseV2(event_type=event_type, scenario=scenario))
    audit(db, request, "V2_INICIALIZAR_HOMOLOGACAO", "HomologationCaseV2", None, "checklist padrão"); db.commit()
    return RedirectResponse("/operacional/homologacao", status_code=303)


@router.post("/homologacao/{case_id}")
def update_homologation(request: Request, case_id: int, status: str = Form(...), protocol: str = Form(""), receipt: str = Form(""), db: Session = Depends(get_db)):
    require_user(request); item = db.get(HomologationCaseV2, case_id)
    if not item:
        raise HTTPException(404)
    if status in {"APROVADO_RESTRITO", "APROVADO_PRODUCAO"} and not protocol.strip():
        raise HTTPException(400, "Aprovação exige protocolo registrado.")
    if status == "APROVADO_PRODUCAO" and not receipt.strip():
        raise HTTPException(400, "Aprovação em produção exige recibo individual.")
    item.status = status; item.protocol = protocol.strip() or None; item.receipt = receipt.strip() or None; audit(db, request, "V2_ATUALIZAR_HOMOLOGACAO", "HomologationCaseV2", str(item.id), status); db.commit()
    return RedirectResponse("/operacional/homologacao", status_code=303)
